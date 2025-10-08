# adcom_contingency.py
import io
import re
from collections import defaultdict, OrderedDict
from datetime import datetime
from typing import Dict, List, Tuple, Optional

import pandas as pd
from openpyxl import load_workbook


# ────────────────────────────────────────────────────────────────────────────────
# Detection regexes
# ────────────────────────────────────────────────────────────────────────────────
DAY_NAMES = r"(Mon|Tue|Tues|Wed|Thu|Thur|Fri|Sat|Sun|Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)"
HEADER_RGX = re.compile(
    rf"{DAY_NAMES}\s*\(\s*\d{{1,2}}/\d{{1,2}}\s*\)\s*[\r\n]+.*?\d{{1,2}}:\d{{2}}\s*[-–]\s*\d{{1,2}}:\d{{2}}\s*(?:[ap]\.?m\.?)?",
    re.IGNORECASE | re.DOTALL,
)
TIME_RANGE_RGX = re.compile(
    r"^\s*(\d{1,2}):(\d{2})\s*[-–]\s*(\d{1,2}):(\d{2})\s*([ap]\.?m\.?)?\s*$",
    re.IGNORECASE,
)
DATE_RGX = re.compile(r"^\s*([A-Za-z]+)\s*\(\s*(\d{1,2})/(\d{1,2})\s*\)\s*$")

DOW_ABBR = {
    "monday": "Mon", "tuesday": "Tue", "wednesday": "Wed", "thursday": "Thu",
    "friday": "Fri", "saturday": "Sat", "sunday": "Sun",
    "mon": "Mon", "tue": "Tue", "tues": "Tue", "wed": "Wed",
    "thu": "Thu", "thur": "Thu", "fri": "Fri", "sat": "Sat", "sun": "Sun",
}

PREFERRED_SHEET_NAMES = (
    "AdCom Availability",
    "Adcom Availability",
    "AdCom_Availability",
    "Adcom_Availability",
    "AdCom",
    "Adcom",
)

# ────────────────────────────────────────────────────────────────────────────────
# Helpers
# ────────────────────────────────────────────────────────────────────────────────
def _normalize_ampm(s: str) -> str:
    s = (s or "").lower().replace(".", "")
    if s.endswith("am") or s.endswith("pm"):
        return s[-2:].upper()
    return ""


def _is_matrix_header_cell(val: str) -> bool:
    val = str(val or "").strip()
    return bool(val and HEADER_RGX.search(val))


def _detect_matrix_sheet(df: pd.DataFrame) -> Tuple[bool, List[int], Dict[int, List[Tuple[int, str]]], int]:
    """
    Return (is_matrix_like, header_row_idxs, header_cells_by_row, strength_score)
      - strength_score = total header-like cells found across all header rows (bigger = stronger match)
    """
    header_rows, header_cells = [], {}
    score = 0
    for i in range(df.shape[0]):
        matches = []
        for j in range(df.shape[1]):
            if _is_matrix_header_cell(df.iat[i, j]):
                matches.append((j, str(df.iat[i, j])))
        if len(matches) >= 3:  # a header row typically has many header cells
            header_rows.append(i)
            header_cells[i] = matches
            score += len(matches)
    return (len(header_rows) > 0), header_rows, header_cells, score


def _parse_header_cell(text: str, year: int) -> Tuple[str, str, str, str]:
    """
    Returns (date_key, display_date_header, start_label, end_label)
      - date_key: "Mon 10/27/2025"
      - display_date_header: "Monday (10/27)"
      - start_label: "8:00 AM"
      - end_label: "9:30 AM"
    """
    parts = [p.strip() for p in str(text).splitlines() if p and str(p).strip()]
    if len(parts) != 2:
        raise ValueError("Header cell not in expected 2-line format")

    m_date = DATE_RGX.match(parts[0])
    m_time = TIME_RANGE_RGX.match(parts[1])
    if not (m_date and m_time):
        raise ValueError("Header cell didn't match date/time patterns")

    dow_raw, mm, dd = m_date.groups()
    h1, m1, h2, m2, ampm = m_time.groups()
    ampm_norm = _normalize_ampm(ampm)

    start_label = f"{int(h1)}:{m1}"
    end_label = f"{int(h2)}:{m2}"
    if ampm_norm:
        start_label = f"{start_label} {ampm_norm}"
        end_label = f"{end_label} {ampm_norm}"

    dow_abbr = DOW_ABBR.get(dow_raw.lower(), dow_raw[:3].title())
    date_key = f"{dow_abbr} {int(mm)}/{int(dd)}/{year}"
    display_date_header = f"{dow_raw.title()} ({int(mm)}/{int(dd)})"
    return date_key, display_date_header, start_label, end_label


def _build_timesheets(
    df: pd.DataFrame,
    header_rows: List[int],
    header_cells: Dict[int, List[Tuple[int, str]]],
    year: int,
):
    """
    timesheets[time_label][date_key] = [names...]
    display_headers_by_date[date_key] = "Monday (10/27)"
    """
    timesheets: Dict[str, "OrderedDict[str, List[str]]"] = defaultdict(OrderedDict)
    display_headers_by_date: Dict[str, str] = {}

    for idx, hr in enumerate(header_rows):
        next_hr = header_rows[idx + 1] if idx + 1 < len(header_rows) else df.shape[0]
        name_rows = range(hr + 1, next_hr)
        for col_idx, header_text in header_cells[hr]:
            try:
                date_key, display_header, start_label, _ = _parse_header_cell(header_text, year)
            except Exception:
                continue
            display_headers_by_date[date_key] = display_header
            names: List[str] = []
            for r in name_rows:
                val = str(df.iat[r, col_idx]).strip()
                if not val or re.fullmatch(r"[-–]+", val):
                    continue
                names.append(val)
            if date_key not in timesheets[start_label]:
                timesheets[start_label][date_key] = []
            timesheets[start_label][date_key].extend(names)

    return timesheets, display_headers_by_date


def _date_sort_key(date_key: str) -> datetime:
    # date_key: "Mon 10/27/2025"
    _, mdY = date_key.split(" ", 1)
    return datetime.strptime(mdY, "%m/%d/%Y")


def _sanitize_sheet(name: str) -> str:
    # Excel forbids : \ / ? * [ ] and >31 chars
    safe = re.sub(r'[:\\/?*\[\]]', '', name).strip()
    return safe[:31]


def _infer_year_from_name(file_like, default_year: int) -> int:
    name = getattr(file_like, "name", "") or ""
    m = re.search(r"(20\d{2})", name)
    return int(m.group(1)) if m else default_year


# ────────────────────────────────────────────────────────────────────────────────
# Public: convert only if matrix-like Adcom sheet found
# ────────────────────────────────────────────────────────────────────────────────
def maybe_convert_adcom_excel(
    file_like,
    example_file_like: Optional[io.BytesIO] = None,
    default_year: int = 2025,
    align_to_file_like: Optional[io.BytesIO] = None,  # NEW: pass the primary workbook (raw)
    header_row_index: int = 3,                         # matches format_xlsx_core header_row
) -> bytes:
    """
    Convert an 'adcom matrix' sheet into time-sheet tabs that format_xlsx_core can parse.

    - Headers at row = header_row_index (0-based) like "Monday (10/27)".
    - Sheet names among {"8 AM","10:30 AM","1:00 PM","3:30 PM","6:00 PM"} which
      format_xlsx_core.format_sheet_time(...) maps to {"8AM","1030AM","1PM","330PM","6PM"}.
    - If align_to_file_like is provided (recommended), the date set/order will match
      the primary workbook's date columns exactly (e.g., "10/27","10/29",...).
    """
    import xlsxwriter
    from xlsxwriter.utility import xl_range

    # --- read original bytes
    xls_bytes = file_like.getvalue() if hasattr(file_like, "getvalue") else file_like.read()
    bio = io.BytesIO(xls_bytes)

    # --- list sheets; if unreadable, pass-through
    try:
        xf = pd.ExcelFile(bio)
        sheet_names = xf.sheet_names
    except Exception:
        return xls_bytes

    # --- find the strongest matrix-like sheet (unchanged logic)
    candidates_priority = []
    lower_map = {s.lower(): s for s in sheet_names}
    for want in PREFERRED_SHEET_NAMES:
        s = lower_map.get(want.lower())
        if s:
            candidates_priority.append(s)
    for s in sheet_names:
        sl = s.lower()
        if ("adcom" in sl) or ("ad com" in sl) or ("adcom_" in sl):
            if s not in candidates_priority:
                candidates_priority.append(s)
    for s in sheet_names:
        if s not in candidates_priority:
            candidates_priority.append(s)

    best = None
    for s in candidates_priority:
        try:
            df = pd.read_excel(io.BytesIO(xls_bytes), sheet_name=s, header=None, dtype=str).fillna("")
        except Exception:
            continue
        is_matrix, hdr_rows, hdr_cells, score = _detect_matrix_sheet(df)
        if is_matrix and (best is None or score > best[3]):
            best = (s, hdr_rows, hdr_cells, score)
    if best is None:
        return xls_bytes

    # --- parse the chosen sheet into (timesheets, display_headers_by_date)
    sheet_name, header_rows, header_cells, _ = best
    df_sheet = pd.read_excel(io.BytesIO(xls_bytes), sheet_name=sheet_name, header=None, dtype=str).fillna("")

    year = _infer_year_from_name(file_like, default_year)
    timesheets, display_headers_by_date = _build_timesheets(df_sheet, header_rows, header_cells, year)
    # timesheets: { "8:00 AM": OrderedDict({ "Mon 10/27/2025": [names...] , ... }) , ... }
    # display_headers_by_date: { "Mon 10/27/2025": "Monday (10/27)", ... }

    # --- (conditional) rewrite 10:00 -> 10:30 ONLY if we will emit a 10:30 tab
    if "10:00 AM" in timesheets:
        if "10:30 AM" not in timesheets:
            timesheets["10:30 AM"] = timesheets["10:00 AM"]
        else:
            for dkey, names in timesheets["10:00 AM"].items():
                timesheets["10:30 AM"].setdefault(dkey, [])
                timesheets["10:30 AM"][dkey].extend(names)
        del timesheets["10:00 AM"]

    # --- decide the final date list/order for headers on every time-sheet
    def _mmdd_from_display(hdr: str) -> str:
        # "Monday (10/27)" -> "10/27"
        m = re.search(r"\((\d{1,2}/\d{1,2})\)", hdr)
        return m.group(1) if m else ""

    # (A) If we can align to the primary workbook, copy its date order exactly
    aligned_mmdd: List[str] = []
    if align_to_file_like is not None:
        try:
            # emulate format_xlsx_core: read every sheet, pick columns that contain "("
            pwb = load_workbook_from_filelike(align_to_file_like)
            seen = []
            for s in pwb.sheet_names:
                try:
                    sheet = pwb.parse(s, header=header_row_index)
                except Exception:
                    continue
                date_columns = [c for c in sheet.columns if isinstance(c, str) and "(" in c]
                for c in date_columns:
                    mmdd = extract_date(c)  # reuses format_xlsx_core behavior
                    if mmdd and mmdd not in seen:
                        seen.append(mmdd)
            aligned_mmdd = seen
        except Exception:
            aligned_mmdd = []

    # (B) If no primary is supplied, derive from the data we parsed (chronological)
    if not aligned_mmdd:
        all_keys = sorted(display_headers_by_date.keys(), key=_date_sort_key)
        aligned_mmdd = []
        for k in all_keys:
            hdr = display_headers_by_date[k]  # "Monday (10/27)"
            mmdd = _mmdd_from_display(hdr)
            if mmdd:
                aligned_mmdd.append(mmdd)

    # Build the display headers in the final order
    mmdd_to_dkey = {}  # "10/27" -> "Mon 10/27/2025"
    for dkey, hdr in display_headers_by_date.items():
        mmdd = _mmdd_from_display(hdr)
        if mmdd and mmdd not in mmdd_to_dkey:
            mmdd_to_dkey[mmdd] = dkey

    final_headers: List[str] = []
    final_date_keys: List[str] = []
    for mmdd in aligned_mmdd:
        dkey = mmdd_to_dkey.get(mmdd)
        if dkey:
            final_headers.append(display_headers_by_date[dkey])  # "Monday (10/27)"
            final_date_keys.append(dkey)

    # if still nothing, bail out pass-through
    if not final_headers:
        return xls_bytes

    # --- target time tabs (these are the ones format_xlsx_core knows)
    desired_sheets = [
        ("8 AM", "8:00 AM"),
        ("10:30 AM", "10:30 AM"),
        ("1:00 PM", "1:00 PM"),
        ("3:30 PM", "3:30 PM"),
        ("6:00 PM", "6:00 PM"),
    ]

    # --- write the normalized workbook
    out = io.BytesIO()
    wb = xlsxwriter.Workbook(out)

    fmt_title    = wb.add_format({"bold": True, "align": "center", "valign": "vcenter"})
    fmt_subtitle = wb.add_format({"bold": True, "align": "center", "valign": "vcenter"})
    fmt_header   = wb.add_format({"bold": True, "align": "center"})
    fmt_name     = wb.add_format({"align": "left"})

    # a light, safe banner
    example_banner = " | ".join(final_headers[:5]) if final_headers else "Adcom Availability"

    for display_name, time_key in desired_sheets:
        ws = wb.add_worksheet(_sanitize_sheet(display_name))

        C = len(final_headers)
        # Merge title row (row 0) and time-window row (row 2) using numeric coords
        ws.merge_range(0, 0, 0, max(0, C - 1), example_banner, fmt_title)
        # Second merged row shows the time window text (optional; purely cosmetic)
        ws.merge_range(2, 0, 2, max(0, C - 1), display_name, fmt_subtitle)

        # Write the header row at header_row_index (default 3 -> Excel row 4)
        for c, hdr in enumerate(final_headers):
            ws.write(header_row_index, c, hdr, fmt_header)

        # Names body: align our internal date_keys to the header order
        date_map = timesheets.get(time_key, OrderedDict())
        col_to_names: List[List[str]] = []
        for dkey in final_date_keys:
            names = list(date_map.get(dkey, []))
            # normalize & de-dup within a column
            seen = set()
            nn = []
            for n in names:
                n = re.sub(r"\s+", " ", str(n)).strip()
                low = n.lower()
                if n and low not in seen:
                    seen.add(low)
                    nn.append(n.title())  # match format_xlsx_core normalization
            col_to_names.append(nn)

        # Write names below the header
        for c, names in enumerate(col_to_names):
            for r, name in enumerate(names, start=header_row_index + 1):
                ws.write(r, c, name, fmt_name)

        # make columns reasonably wide
        ws.set_column(0, max(0, C - 1), 20)

    wb.close()
    out.seek(0)
    return out.read()

