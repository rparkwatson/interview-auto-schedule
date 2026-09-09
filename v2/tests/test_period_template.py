from __future__ import annotations

from datetime import datetime, time, timedelta
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
import pytest

from interview_scheduler_v2.domain import Slot
from interview_scheduler_v2.io import (
    PeriodTemplateError,
    build_interview_period_template,
    parse_completed_interview_period_template,
    period_template_filename,
)


def candidate_slots() -> tuple[Slot, ...]:
    zone = ZoneInfo("America/New_York")
    first = datetime(2026, 2, 24, 8, 0, tzinfo=zone)
    second = datetime(2026, 2, 24, 10, 30, tzinfo=zone)
    return (
        Slot("20260224-0800", first, first + timedelta(minutes=90), 0, 0),
        Slot("20260224-1030", second, second + timedelta(minutes=90), 0, 0),
    )


def completed_template_bytes(content: bytes) -> bytes:
    workbook = load_workbook(BytesIO(content))
    sheet = workbook["Interview Periods"]
    sheet["F5"] = 3
    sheet["F6"] = 2
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_period_template_is_fillable_protected_and_round_trips_counts():
    slots = candidate_slots()
    content = build_interview_period_template(
        slots,
        scenario="Winter 2026 / Round 2",
        student_available={slots[0].id: 4, slots[1].id: 2},
        adcom_available={slots[0].id: 3, slots[1].id: 1},
    )
    workbook = load_workbook(BytesIO(content), data_only=False)
    assert workbook.sheetnames == ["Interview Periods", "Template Metadata"]
    assert workbook["Template Metadata"].sheet_state == "hidden"
    sheet = workbook["Interview Periods"]
    assert sheet.protection.sheet
    assert sheet["A5"].value == slots[0].id
    assert sheet["B5"].is_date
    assert [cell.value for cell in sheet[4]] == [
        "Period ID",
        "Date",
        "Start time",
        "End time",
        "Interview period",
        "Interviews possible",
        "Student available",
        "Adcom available",
    ]
    assert sheet["G5"].value == 4
    assert sheet["H5"].value == 3
    assert not sheet["F5"].protection.locked
    assert sheet["G5"].protection.locked
    assert sheet["E5"].protection.locked

    parsed = parse_completed_interview_period_template(
        completed_template_bytes(content),
        expected_slots=slots,
        year=2026,
    )
    assert [(item.capacity, item.target) for item in parsed.slots] == [(3, 3), (2, 2)]


def test_period_template_rejects_blank_required_counts():
    slots = candidate_slots()
    content = build_interview_period_template(slots, scenario="Winter 2026")
    with pytest.raises(PeriodTemplateError, match="Complete every"):
        parse_completed_interview_period_template(
            content,
            expected_slots=slots,
            year=2026,
        )


def test_period_template_rejects_a_file_from_another_availability_setup():
    slots = candidate_slots()
    content = completed_template_bytes(
        build_interview_period_template(slots, scenario="Winter 2026")
    )
    changed_start = slots[1].start + timedelta(days=1)
    changed_slots = (
        slots[0],
        Slot(
            "20260225-1030",
            changed_start,
            changed_start + timedelta(minutes=90),
            0,
            0,
        ),
    )
    with pytest.raises(PeriodTemplateError, match="different availability setup"):
        parse_completed_interview_period_template(
            content,
            expected_slots=changed_slots,
            year=2026,
        )


def test_period_template_rejects_changed_dates_or_times():
    slots = candidate_slots()
    content = completed_template_bytes(
        build_interview_period_template(slots, scenario="Winter 2026")
    )
    workbook = load_workbook(BytesIO(content))
    workbook["Interview Periods"]["C5"] = time(9, 0)
    output = BytesIO()
    workbook.save(output)
    with pytest.raises(PeriodTemplateError, match="dates or times were changed"):
        parse_completed_interview_period_template(
            output.getvalue(),
            expected_slots=slots,
            year=2026,
        )


def test_period_template_filename_is_scenario_specific():
    assert (
        period_template_filename("Winter 2026 / Round 2")
        == "Winter_2026_Round_2_interview_periods.xlsx"
    )
