from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from interview_scheduler_v2 import (
    GroupPolicy,
    Interviewer,
    InterviewerGroup,
    SchedulerConfig,
    SchedulingProblem,
    Slot,
)
from interview_scheduler_v2.optimization import (
    Assignment,
    SlotSummary,
    SolveResult,
    SolveStatus,
    solve,
)
from interview_scheduler_v2.reporting import (
    build_simplified_schedule_workbook,
    build_workbook,
    scenario_filename,
    simplified_schedule_filename,
)
from interview_scheduler_v2.reporting.excel import (
    REQUIRED_SHEETS,
    SIMPLIFIED_SCHEDULE_SHEET,
)


def test_required_workbook_is_generated_in_memory_with_assignment_contract():
    start = datetime(2026, 3, 1, 8, tzinfo=ZoneInfo("America/New_York"))
    slot = Slot("s1", start, start + timedelta(minutes=90), 1, 1)
    person = Interviewer.create(
        name="Student One",
        group=InterviewerGroup.STUDENT,
        available_slot_ids=["s1"],
    )
    problem = SchedulingProblem((person,), (slot,))
    config = SchedulerConfig(
        group_policies={
            InterviewerGroup.STUDENT: GroupPolicy(1, 1, 1, 1),
            InterviewerGroup.ADCOM: GroupPolicy(0, 0, 1, 1),
        },
        time_limit_seconds=5,
    )
    result = solve(problem, scenario="Winter 2026", config=config)
    content = build_workbook(result, problem, config)

    assert content.startswith(b"PK")
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    assert workbook.sheetnames == list(REQUIRED_SHEETS)
    assignment_sheet = workbook["Assignments"]
    assert [cell.value for cell in assignment_sheet[1]] == [
        "Scenario",
        "Slot ID",
        "Start",
        "End",
        "Interviewer ID",
        "Interviewer Name",
        "Group",
        "Locked",
        "Preference",
    ]
    assert assignment_sheet["F2"].value == "Student One"
    assert assignment_sheet["G2"].value == "Student Interviewer"

    simplified_content = build_simplified_schedule_workbook(result)
    simplified = load_workbook(
        BytesIO(simplified_content),
        read_only=True,
        data_only=False,
    )[SIMPLIFIED_SCHEDULE_SHEET]
    assert simplified["C2"].value == assignment_sheet["F2"].value


def test_filename_is_scenario_specific_and_sanitized():
    value = scenario_filename(
        "Winter 2026 / Round 2",
        generated_at=datetime(2026, 8, 28, 14, 30),
    )
    assert value == "Winter_2026_Round_2_interview_schedule_20260828_143000.xlsx"


def test_simplified_schedule_has_one_sorted_row_per_slot_and_capacity_columns():
    zone = ZoneInfo("America/New_York")
    early = datetime(2026, 3, 1, 8, tzinfo=zone)
    later = datetime(2026, 3, 1, 10, 30, tzinfo=zone)
    next_day = datetime(2026, 3, 2, 8, tzinfo=zone)
    result = SolveResult(
        status=SolveStatus.OPTIMAL,
        scenario="Winter 2026",
        assignments=(
            Assignment(
                "Winter 2026",
                "late",
                later,
                later + timedelta(minutes=90),
                "ADC-1",
                "Adcom Later",
                InterviewerGroup.ADCOM,
                False,
                1,
            ),
            Assignment(
                "Winter 2026",
                "early",
                early,
                early + timedelta(minutes=90),
                "STU-1",
                "Student Early",
                InterviewerGroup.STUDENT,
                False,
                1,
            ),
            Assignment(
                "Winter 2026",
                "early",
                early,
                early + timedelta(minutes=90),
                "ADC-2",
                "Adcom Early",
                InterviewerGroup.ADCOM,
                False,
                1,
            ),
        ),
        slot_summaries=(
            SlotSummary("late", later, later + timedelta(minutes=90), 3, 3, 1, 2, 2, 0, 1, None, None),
            SlotSummary("next", next_day, next_day + timedelta(minutes=90), 1, 1, 0, 1, 1, 0, 0, None, None),
            SlotSummary("early", early, early + timedelta(minutes=90), 2, 2, 2, 0, 0, 1, 1, None, None),
        ),
    )

    content = build_simplified_schedule_workbook(result)

    assert content.startswith(b"PK")
    workbook = load_workbook(BytesIO(content), data_only=False)
    assert workbook.sheetnames == [SIMPLIFIED_SCHEDULE_SHEET]
    worksheet = workbook[SIMPLIFIED_SCHEDULE_SHEET]
    assert [cell.value for cell in worksheet[1]] == [
        "Slot Date",
        "Start Time",
        "Group A",
        "Group B",
        "Group C",
    ]
    assert worksheet["A2"].value == early.replace(tzinfo=None)
    assert worksheet["A2"].number_format == "mm/dd"
    assert worksheet["B2"].value == early.replace(tzinfo=None)
    assert worksheet["B2"].number_format == "h:mm AM/PM"
    assert [worksheet.cell(2, column).value for column in range(3, 6)] == [
        "Student Early",
        "Adcom Early",
        None,
    ]
    assert worksheet["C3"].value == "Adcom Later"
    assert worksheet["A4"].value == next_day.replace(tzinfo=None)
    assert worksheet["D4"].value is None
    assert worksheet["D4"].fill.fgColor.rgb == "FFE7E6E6"
    assert worksheet.freeze_panes == "C2"


def test_simplified_schedule_supports_group_columns_beyond_z():
    start = datetime(2026, 3, 1, 8, tzinfo=ZoneInfo("America/New_York"))
    result = SolveResult(
        status=SolveStatus.OPTIMAL,
        scenario="Large period",
        slot_summaries=(
            SlotSummary("s1", start, start + timedelta(minutes=90), 28, 28, 0, 28, 28, 0, 0, None, None),
        ),
    )

    workbook = load_workbook(
        BytesIO(build_simplified_schedule_workbook(result)),
        read_only=True,
    )
    headers = [cell.value for cell in workbook[SIMPLIFIED_SCHEDULE_SHEET][1]]
    assert headers[-3:] == ["Group Z", "Group AA", "Group AB"]


def test_simplified_filename_is_scenario_specific_and_sanitized():
    value = simplified_schedule_filename(
        "Winter 2026 / Round 2",
        generated_at=datetime(2026, 8, 28, 14, 30),
    )
    assert value == "Winter_2026_Round_2_simplified_schedule_20260828_143000.xlsx"
