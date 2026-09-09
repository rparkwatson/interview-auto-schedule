from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from interview_scheduler_v2.domain import InterviewerGroup
from interview_scheduler_v2.io import import_campaign, prepare_campaign_from_availability
from interview_scheduler_v2.io.source_parsers import (
    parse_adcom_availability,
    parse_student_availability,
    parse_student_schedule,
    parse_time_slot_workbook,
    reconcile_availability,
)


def workbook_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def slot_source() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Time Slots"
    sheet.append(["Date_Time", "Max_Slot", "Target"])
    sheet.append(["2/24 - 800 AM", 2, 1])
    sheet.append(["2/24 - 1030 AM", 3, 3])
    return workbook_bytes(workbook)


def explicit_slot_source() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Slot ID", "Date", "Start Time", "End Time", "Capacity"])
    sheet.append(["ROUND2-001", "2026-02-24", "8:00 AM", "9:30 AM", 4])
    return workbook_bytes(workbook)


def student_source() -> bytes:
    workbook = Workbook()
    example = workbook.active
    example.title = "EXAMPLE"
    example.append(["Names"])
    example.append(["Wrong Person"])

    roster = workbook.create_sheet("AF Names")
    roster.append(["AF Names"])
    roster.append(["Student One"])
    roster.append(["Student Zero"])

    sheet = workbook.create_sheet("800 AM")
    sheet.cell(3, 1, "8:00 am - 9:30 am")
    sheet.cell(4, 2, "Tuesday (2/24)")
    sheet.cell(4, 3, "Wednesday (2/25)")
    sheet.cell(5, 2, "Student   One")

    ignored = workbook.create_sheet("IGNORE Robin")
    ignored.cell(3, 1, "8:00 am - 9:30 am")
    ignored.cell(4, 2, "Tuesday (2/24)")
    ignored.cell(4, 3, "Wednesday (2/25)")
    ignored.cell(5, 2, "Wrong Person")
    return workbook_bytes(workbook)


def adcom_source() -> bytes:
    workbook = Workbook()
    example = workbook.active
    example.title = "EXAMPLE"
    example.cell(1, 1, "Tuesday (2/24) 8:00 am - 9:30 am")
    example.cell(2, 1, "Wrong Person")

    availability = workbook.create_sheet("AdCom Availability")
    availability.cell(2, 1, "Tuesday (2/24) 8:00 am - 9:30 am")
    availability.cell(2, 2, "Tuesday (2/24) 10:00 am - 11:30 am")
    availability.cell(3, 1, "Adcom One")
    availability.cell(3, 2, "Adcom One")

    links = workbook.create_sheet("AdCom Zoom Links")
    links.cell(1, 1, "Tuesday (2/24) 8:00 am - 9:30 am")
    links.cell(2, 1, "credential-like-value")
    return workbook_bytes(workbook)


def student_single_empty_period_source() -> bytes:
    workbook = Workbook()
    roster = workbook.active
    roster.title = "AF Names"
    roster.append(["AF Names"])
    roster.append(["Student One"])
    sheet = workbook.create_sheet("800 AM")
    sheet.cell(3, 1, "8:00 am - 9:30 am")
    sheet.cell(4, 2, "Tuesday (2/24)")
    return workbook_bytes(workbook)


def test_time_slots_are_authoritative_and_target_defaults_are_supported():
    parsed = parse_time_slot_workbook(slot_source(), year=2026)
    assert [slot.id for slot in parsed.slots] == ["20260224-0800", "20260224-1030"]
    assert parsed.slots[0].capacity == 2
    assert parsed.slots[0].target == 1
    assert parsed.slots[0].end.hour == 9
    assert parsed.slots[0].end.minute == 30
    assert str(parsed.slots[0].start.tzinfo) == "America/New_York"


def test_explicit_date_start_end_and_slot_id_columns_are_supported():
    parsed = parse_time_slot_workbook(explicit_slot_source(), year=2026)
    assert len(parsed.slots) == 1
    assert parsed.slots[0].id == "ROUND2-001"
    assert parsed.slots[0].capacity == parsed.slots[0].target == 4
    assert parsed.slots[0].end.hour == 9
    assert parsed.slots[0].end.minute == 30


def test_student_parser_uses_time_sheets_and_preserves_zero_availability_roster():
    parsed = parse_student_availability(student_source(), year=2026)
    assert parsed.names == ["Student One", "Student Zero"]
    assert {entry.interviewer_name for entry in parsed.entries} == {"Student One"}
    assert {entry.source_sheet for entry in parsed.entries} == {"800 AM"}


def test_student_headers_create_periods_even_when_no_one_is_available():
    parsed = parse_student_schedule(student_source(), year=2026)
    assert [slot.id for slot in parsed.slot_set.slots] == [
        "20260224-0800",
        "20260225-0800",
    ]
    assert all(slot.capacity == slot.target == 0 for slot in parsed.slot_set.slots)
    assert {
        entry.slot_id
        for entry in reconcile_availability(
            parsed.availability,
            parsed.slot_set,
        ).entries
    } == {"20260224-0800"}


def test_single_student_period_is_retained_with_zero_availability():
    parsed = parse_student_schedule(student_single_empty_period_source(), year=2026)
    assert parsed.availability.names == ["Student One"]
    assert parsed.availability.entries == []
    assert [slot.id for slot in parsed.slot_set.slots] == ["20260224-0800"]


def test_adcom_parser_excludes_example_and_zoom_sheets_and_reconciles_30_minutes():
    slots = parse_time_slot_workbook(slot_source(), year=2026)
    parsed = parse_adcom_availability(adcom_source(), year=2026)
    reconciled = reconcile_availability(parsed, slots)
    assert parsed.names == ["Adcom One"]
    assert {entry.source_sheet for entry in parsed.entries} == {"AdCom Availability"}
    assert {entry.slot_id for entry in reconciled.entries} == {
        "20260224-0800",
        "20260224-1030",
    }
    assert len([n for n in reconciled.notices if n.code == "slot_time_reconciled"]) == 1


def test_campaign_builds_single_people_and_stable_group_ids():
    imported = import_campaign(
        student_workbook=student_source(),
        adcom_workbook=adcom_source(),
        slot_workbook=slot_source(),
        year=2026,
    )
    assert len(imported.problem.interviewers) == 3
    student_ids = {
        person.id
        for person in imported.problem.interviewers
        if person.group is InterviewerGroup.STUDENT
    }
    adcom_ids = {
        person.id
        for person in imported.problem.interviewers
        if person.group is InterviewerGroup.ADCOM
    }
    assert all(value.startswith("STU-") for value in student_ids)
    assert all(value.startswith("ADC-") for value in adcom_ids)
    assert next(
        person
        for person in imported.problem.interviewers
        if person.name == "Student Zero"
    ).available_slot_ids == frozenset()


def test_campaign_can_be_prepared_without_a_separate_slot_workbook():
    imported = prepare_campaign_from_availability(
        student_workbook=student_source(),
        adcom_workbook=adcom_source(),
        year=2026,
    )
    assert imported.periods_need_configuration
    assert [slot.id for slot in imported.problem.slots] == [
        "20260224-0800",
        "20260225-0800",
    ]
    assert all(slot.capacity == 0 for slot in imported.problem.slots)
    adcom = next(
        person
        for person in imported.problem.interviewers
        if person.group is InterviewerGroup.ADCOM
    )
    assert adcom.available_slot_ids == frozenset({"20260224-0800"})
