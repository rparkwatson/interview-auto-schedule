from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from io import BytesIO
from pathlib import Path
import re
from typing import BinaryIO, Iterable
import warnings
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


WorkbookSource = str | Path | bytes | bytearray | BinaryIO


@dataclass(frozen=True)
class ImportNotice:
    severity: str
    code: str
    message: str
    source: str | None = None


@dataclass(frozen=True)
class ParsedSlot:
    id: str
    start: datetime
    end: datetime
    target: int
    capacity: int
    source_label: str


@dataclass
class ParsedSlotSet:
    slots: list[ParsedSlot]
    notices: list[ImportNotice] = field(default_factory=list)


@dataclass(frozen=True)
class ParsedAvailabilityEntry:
    interviewer_name: str
    source_date: date
    source_start: time
    source_end: time
    source_label: str
    source_sheet: str
    slot_id: str | None = None


@dataclass
class ParsedAvailability:
    group: str
    entries: list[ParsedAvailabilityEntry]
    names: list[str]
    notices: list[ImportNotice] = field(default_factory=list)


@dataclass
class ParsedStudentSchedule:
    """Student availability plus every date/time period found in its headers."""

    availability: ParsedAvailability
    slot_set: ParsedSlotSet


_DATE_IN_PARENS = re.compile(r"\((\d{1,2})/(\d{1,2})\)")
_TIME_TOKEN = re.compile(r"(\d{1,2})(?::?(\d{2}))?\s*([ap])?\.?m?\.?", re.IGNORECASE)
_TIME_RANGE = re.compile(
    r"(\d{1,2})(?::(\d{2}))?\s*([ap])?\.?m?\.?\s*[-–]\s*"
    r"(\d{1,2})(?::(\d{2}))?\s*([ap])?\.?m?\.?",
    re.IGNORECASE,
)
_DATE_TIME_SHORT = re.compile(
    r"^\s*(\d{1,2})/(\d{1,2})\s*[-–]\s*([0-9: ]+\s*(?:AM|PM))\s*$",
    re.IGNORECASE,
)


def _workbook_input(source: WorkbookSource):
    if isinstance(source, (bytes, bytearray)):
        return BytesIO(source)
    return source


def _load_source_workbook(source: WorkbookSource):
    # The supplied templates contain legacy Excel data-validation extensions.
    # We only read values and never save the source workbook, so openpyxl's
    # unsupported-extension warning is noise rather than data loss in v2.
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported.*",
            category=UserWarning,
        )
        return load_workbook(
            _workbook_input(source),
            read_only=False,
            data_only=True,
        )


def _clean_name(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _unique_names(entries: Iterable[ParsedAvailabilityEntry]) -> list[str]:
    seen: set[str] = set()
    names: list[str] = []
    for entry in entries:
        key = entry.interviewer_name.casefold()
        if key and key not in seen:
            seen.add(key)
            names.append(entry.interviewer_name)
    return names


def _meridiem_hour(hour: int, meridiem: str) -> int:
    hour %= 12
    return hour + (12 if meridiem.lower() == "p" else 0)


def _parse_time_range(text: str) -> tuple[time, time] | None:
    match = _TIME_RANGE.search(text.replace("\n", " "))
    if not match:
        return None

    start_hour = int(match.group(1))
    start_minute = int(match.group(2) or 0)
    start_meridiem = match.group(3)
    end_hour = int(match.group(4))
    end_minute = int(match.group(5) or 0)
    end_meridiem = match.group(6)

    if not end_meridiem and not start_meridiem:
        return None
    if not end_meridiem:
        end_meridiem = start_meridiem
    if not start_meridiem:
        start_meridiem = end_meridiem
        # A range such as 11:30 - 1:00 pm crosses noon.
        if end_meridiem.lower() == "p" and start_hour > end_hour:
            start_meridiem = "a"

    return (
        time(_meridiem_hour(start_hour, start_meridiem), start_minute),
        time(_meridiem_hour(end_hour, end_meridiem), end_minute),
    )


def _parse_short_time(text: str) -> time:
    normalized = re.sub(r"\s+", "", text).upper().replace(":", "")
    match = re.fullmatch(r"(\d{1,4})(AM|PM)", normalized)
    if not match:
        raise ValueError(f"Unrecognized time label: {text!r}")
    digits, meridiem = match.groups()
    if len(digits) <= 2:
        hour, minute = int(digits), 0
    else:
        hour, minute = int(digits[:-2]), int(digits[-2:])
    if not 1 <= hour <= 12 or not 0 <= minute <= 59:
        raise ValueError(f"Invalid time label: {text!r}")
    return time(_meridiem_hour(hour, meridiem[0]), minute)


def _aware_datetime(day: date, value: time, timezone_name: str) -> datetime:
    return datetime.combine(day, value, tzinfo=ZoneInfo(timezone_name))


def _slot_id(start: datetime) -> str:
    return start.strftime("%Y%m%d-%H%M")


def _format_clock(value: datetime | time) -> str:
    return value.strftime("%I:%M %p").lstrip("0")


def parse_time_slot_workbook(
    source: WorkbookSource,
    *,
    year: int,
    timezone_name: str = "America/New_York",
    default_duration_minutes: int = 90,
) -> ParsedSlotSet:
    """Parse the user-maintained slot target/capacity workbook.

    The example contract uses ``Date_Time`` and ``Max_Slot``. V2 also accepts
    optional ``Target``, ``Capacity``, and ``End`` columns.
    """

    # Normal mode is intentional: the supplied workbooks have inflated stored
    # dimensions and read-only random access becomes prohibitively slow.
    workbook = _load_source_workbook(source)
    notices: list[ImportNotice] = []
    slots: list[ParsedSlot] = []

    header_aliases = {
        "date_time": {"date_time", "datetime", "slot"},
        "date": {"date", "slot_date"},
        "start": {"start", "start_time", "time"},
        "slot_id": {"slot_id", "period_id", "id"},
        "capacity": {
            "max_slot",
            "capacity",
            "max_capacity",
            "maximum",
            "interviews_possible",
        },
        "target": {
            "target",
            "target_assignments",
            "desired",
            "interviews_to_schedule",
            "interviews_to_schedule_optional",
        },
        "end": {"end", "end_time"},
    }

    table_found = False
    for worksheet in workbook.worksheets:
        header_row = None
        columns: dict[str, int] = {}
        for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=min(20, worksheet.max_row), values_only=True), start=1):
            normalized = {
                re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_"): index
                for index, value in enumerate(row)
                if value is not None
            }
            candidate: dict[str, int] = {}
            for canonical, aliases in header_aliases.items():
                for alias in aliases:
                    if alias in normalized:
                        candidate[canonical] = normalized[alias]
                        break
            has_time = "date_time" in candidate or {
                "date",
                "start",
            }.issubset(candidate)
            if has_time and "capacity" in candidate:
                header_row, columns = row_index, candidate
                break
        if header_row is None:
            continue

        table_found = True
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if "date_time" in columns:
                raw_label = (
                    row[columns["date_time"]]
                    if columns["date_time"] < len(row)
                    else None
                )
                if raw_label is None or not str(raw_label).strip():
                    continue
                label = str(raw_label).strip()
                if isinstance(raw_label, datetime):
                    start = _aware_datetime(
                        raw_label.date(),
                        raw_label.time(),
                        timezone_name,
                    )
                else:
                    match = _DATE_TIME_SHORT.match(label)
                    if not match:
                        notices.append(ImportNotice("error", "invalid_slot_label", f"Could not parse slot label {label!r}.", f"{worksheet.title}!row {row_number}"))
                        continue
                    month, day = int(match.group(1)), int(match.group(2))
                    start_time = _parse_short_time(match.group(3))
                    start = _aware_datetime(
                        date(year, month, day),
                        start_time,
                        timezone_name,
                    )
            else:
                raw_date = row[columns["date"]] if columns["date"] < len(row) else None
                raw_start = row[columns["start"]] if columns["start"] < len(row) else None
                if raw_date in (None, "") and raw_start in (None, ""):
                    continue
                try:
                    if isinstance(raw_date, datetime):
                        source_date = raw_date.date()
                    elif isinstance(raw_date, date):
                        source_date = raw_date
                    else:
                        date_text = str(raw_date).strip()
                        month_day = re.fullmatch(r"(\d{1,2})/(\d{1,2})", date_text)
                        source_date = (
                            date(year, int(month_day.group(1)), int(month_day.group(2)))
                            if month_day
                            else date.fromisoformat(date_text)
                        )
                    if isinstance(raw_start, datetime):
                        start_time = raw_start.time()
                    elif isinstance(raw_start, time):
                        start_time = raw_start
                    else:
                        start_time = _parse_short_time(str(raw_start))
                    start = _aware_datetime(source_date, start_time, timezone_name)
                except (TypeError, ValueError) as exc:
                    notices.append(ImportNotice("error", "invalid_slot_label", f"Could not parse date/start values: {exc}", f"{worksheet.title}!row {row_number}"))
                    continue
                label = f"{source_date.isoformat()} {_format_clock(start)}"

            end = start + timedelta(minutes=default_duration_minutes)
            if "end" in columns and columns["end"] < len(row) and row[columns["end"]] not in (None, ""):
                raw_end = row[columns["end"]]
                if isinstance(raw_end, datetime):
                    end = _aware_datetime(start.date(), raw_end.time(), timezone_name)
                elif isinstance(raw_end, time):
                    end = _aware_datetime(start.date(), raw_end, timezone_name)
                else:
                    end = _aware_datetime(start.date(), _parse_short_time(str(raw_end)), timezone_name)

            try:
                capacity = int(row[columns["capacity"]])
            except (TypeError, ValueError, IndexError):
                notices.append(ImportNotice("error", "invalid_capacity", f"Capacity must be an integer for {label!r}.", f"{worksheet.title}!row {row_number}"))
                continue
            if capacity < 0:
                notices.append(ImportNotice("error", "negative_capacity", f"Capacity cannot be negative for {label!r}.", f"{worksheet.title}!row {row_number}"))
                continue

            target = capacity
            if "target" in columns and columns["target"] < len(row) and row[columns["target"]] not in (None, ""):
                try:
                    target = int(row[columns["target"]])
                except (TypeError, ValueError):
                    notices.append(ImportNotice("error", "invalid_target", f"Target must be an integer for {label!r}.", f"{worksheet.title}!row {row_number}"))
                    continue
            if target < 0 or target > capacity:
                notices.append(ImportNotice("error", "target_outside_capacity", f"Target {target} must be between zero and capacity {capacity} for {label!r}.", f"{worksheet.title}!row {row_number}"))
                continue

            explicit_slot_id = None
            if "slot_id" in columns and columns["slot_id"] < len(row):
                raw_slot_id = row[columns["slot_id"]]
                if raw_slot_id is not None and str(raw_slot_id).strip():
                    explicit_slot_id = str(raw_slot_id).strip()
            slots.append(ParsedSlot(explicit_slot_id or _slot_id(start), start, end, target, capacity, label))

    if not table_found:
        raise ValueError("No slot table with Date_Time and Max_Slot/Capacity columns was found.")

    seen: set[str] = set()
    duplicates: set[str] = set()
    for slot in slots:
        if slot.id in seen:
            duplicates.add(slot.id)
        seen.add(slot.id)
    for duplicate in sorted(duplicates):
        notices.append(ImportNotice("error", "duplicate_slot", f"Slot {duplicate} occurs more than once."))

    slots.sort(key=lambda item: item.start)
    workbook.close()
    return ParsedSlotSet(slots=slots, notices=notices)


def parse_student_schedule(
    source: WorkbookSource,
    *,
    year: int,
    timezone_name: str = "America/New_York",
) -> ParsedStudentSchedule:
    """Read Student availability and derive its complete candidate-period list.

    Periods come from the date headers and time range on each Student time
    worksheet. They are retained even when no interviewer entered availability
    for that particular date/time combination.
    """

    workbook = _load_source_workbook(source)
    entries: list[ParsedAvailabilityEntry] = []
    notices: list[ImportNotice] = []
    period_notices: list[ImportNotice] = []
    periods_by_id: dict[str, ParsedSlot] = {}
    normalized_whitespace = 0

    roster: list[str] = []
    roster_candidates = []
    for worksheet in workbook.worksheets:
        normalized_title = re.sub(
            r"[^a-z0-9]+", " ", worksheet.title.casefold()
        ).strip()
        if any(token in normalized_title for token in ("example", "ignore")):
            continue
        first_header = _clean_name(worksheet.cell(1, 1).value).casefold()
        if "name" not in first_header:
            continue
        preferred = 1 if normalized_title in {"af names", "student names"} else 0
        roster_candidates.append((preferred, worksheet))

    for _, worksheet in sorted(
        roster_candidates,
        key=lambda item: item[0],
        reverse=True,
    ):
        seen_roster: set[str] = set()
        for row_index in range(2, worksheet.max_row + 1):
            name = _clean_name(worksheet.cell(row_index, 1).value)
            if not name:
                continue
            key = name.casefold()
            if key in seen_roster:
                notices.append(ImportNotice("error", "duplicate_student_roster_name", f"Student roster contains duplicate name {name!r}.", worksheet.title))
                continue
            seen_roster.add(key)
            roster.append(name)
        if roster:
            break

    for worksheet in workbook.worksheets:
        normalized_title = re.sub(r"\s+", " ", worksheet.title.strip().casefold())
        # Student source sheets use a time as the sheet role. This explicit role
        # rule prevents example/helper sheets from being mistaken for inputs.
        if not re.fullmatch(r"\d{1,4}\s*(?:am|pm)", normalized_title):
            continue
        best_row: int | None = None
        best_headers: list[tuple[int, date]] = []
        for row_index in range(1, min(15, worksheet.max_row) + 1):
            headers: list[tuple[int, date]] = []
            for column_index in range(1, worksheet.max_column + 1):
                value = worksheet.cell(row_index, column_index).value
                match = _DATE_IN_PARENS.search(str(value or ""))
                if match:
                    headers.append((column_index, date(year, int(match.group(1)), int(match.group(2)))))
            if len(headers) > len(best_headers):
                best_row, best_headers = row_index, headers

        if best_row is None or not best_headers:
            continue

        range_value = None
        for row_index in range(max(1, best_row - 3), best_row):
            for column_index in range(1, min(3, worksheet.max_column) + 1):
                candidate = str(worksheet.cell(row_index, column_index).value or "")
                if _parse_time_range(candidate):
                    range_value = candidate
                    break
            if range_value:
                break
        parsed_range = _parse_time_range(range_value or "")
        if not parsed_range:
            notices.append(ImportNotice("warning", "sheet_ignored_no_time_range", f"Ignored sheet {worksheet.title!r}: date headers were found but no time range was recognized.", worksheet.title))
            continue
        start_time, end_time = parsed_range

        for column_index, source_date in best_headers:
            header_label = str(worksheet.cell(best_row, column_index).value or "").strip()
            start = _aware_datetime(source_date, start_time, timezone_name)
            end = _aware_datetime(source_date, end_time, timezone_name)
            if end <= start:
                end += timedelta(days=1)
            slot_id = _slot_id(start)
            period = ParsedSlot(
                id=slot_id,
                start=start,
                end=end,
                target=0,
                capacity=0,
                source_label=header_label,
            )
            existing = periods_by_id.get(slot_id)
            if existing is None:
                periods_by_id[slot_id] = period
            elif existing.start != period.start or existing.end != period.end:
                period_notices.append(
                    ImportNotice(
                        "error",
                        "conflicting_student_period",
                        (
                            f"Student period {slot_id} has more than one time range "
                            "and cannot be used until the source workbook is corrected."
                        ),
                        worksheet.title,
                    )
                )
            for row_index in range(best_row + 1, worksheet.max_row + 1):
                raw_value = worksheet.cell(row_index, column_index).value
                if raw_value is None or not str(raw_value).strip():
                    continue
                raw_name = str(raw_value)
                name = _clean_name(raw_name)
                if name != raw_name:
                    normalized_whitespace += 1
                entries.append(ParsedAvailabilityEntry(name, source_date, start_time, end_time, header_label, worksheet.title))

    if normalized_whitespace:
        notices.append(ImportNotice("info", "name_whitespace_normalized", f"Normalized whitespace in {normalized_whitespace} Student availability cells."))
    if not periods_by_id:
        raise ValueError(
            "No Student availability sheets with date columns and time ranges were found."
        )
    discovered = _unique_names(entries)
    if roster:
        roster_keys = {name.casefold() for name in roster}
        unknown = [name for name in discovered if name.casefold() not in roster_keys]
        for name in unknown:
            notices.append(ImportNotice("error", "student_name_not_in_roster", f"Availability contains Student name {name!r} that is not in the roster."))
        names = roster + unknown
    else:
        names = discovered
        notices.append(ImportNotice("warning", "student_roster_missing", "No explicit Student roster was found; people with zero availability cannot be represented."))
    workbook.close()
    availability = ParsedAvailability("student", entries, names, notices)
    slot_set = ParsedSlotSet(
        slots=sorted(periods_by_id.values(), key=lambda item: item.start),
        notices=period_notices,
    )
    return ParsedStudentSchedule(availability=availability, slot_set=slot_set)


def parse_student_availability(source: WorkbookSource, *, year: int) -> ParsedAvailability:
    """Compatibility wrapper for callers that only need availability records."""

    return parse_student_schedule(source, year=year).availability


def parse_adcom_availability(source: WorkbookSource, *, year: int) -> ParsedAvailability:
    workbook = _load_source_workbook(source)
    candidate_worksheets = []
    for worksheet in workbook.worksheets:
        normalized_title = re.sub(
            r"[^a-z0-9]+", " ", worksheet.title.casefold()
        ).strip()
        # Never inspect credential/helper/example sheets as availability data.
        if any(token in normalized_title for token in ("zoom", "link", "example")):
            continue
        header_cells = 0
        for row in worksheet.iter_rows(values_only=True):
            for value in row:
                text = str(value or "")
                if _DATE_IN_PARENS.search(text) and _parse_time_range(text):
                    header_cells += 1
        if header_cells:
            preferred = 1 if normalized_title == "adcom availability" else 0
            candidate_worksheets.append((preferred, header_cells, worksheet))
    if not candidate_worksheets:
        raise ValueError("No Adcom availability matrix containing date/time headers was found.")

    _, _, worksheet = max(
        candidate_worksheets,
        key=lambda item: (item[0], item[1]),
    )
    header_rows: dict[int, list[tuple[int, date, time, time, str]]] = {}
    for row_index in range(1, worksheet.max_row + 1):
        headers = []
        for column_index in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row_index, column_index).value
            text = str(value or "")
            date_match = _DATE_IN_PARENS.search(text)
            parsed_range = _parse_time_range(text)
            if date_match and parsed_range:
                source_date = date(year, int(date_match.group(1)), int(date_match.group(2)))
                headers.append((column_index, source_date, parsed_range[0], parsed_range[1], text.strip()))
        if headers:
            header_rows[row_index] = headers

    entries: list[ParsedAvailabilityEntry] = []
    notices: list[ImportNotice] = []
    sorted_header_rows = sorted(header_rows)
    for position, row_index in enumerate(sorted_header_rows):
        next_row = sorted_header_rows[position + 1] if position + 1 < len(sorted_header_rows) else worksheet.max_row + 1
        for column_index, source_date, start_time, end_time, label in header_rows[row_index]:
            for name_row in range(row_index + 1, next_row):
                raw_name = worksheet.cell(name_row, column_index).value
                if raw_name is None or not str(raw_name).strip():
                    continue
                name = _clean_name(raw_name)
                entries.append(ParsedAvailabilityEntry(name, source_date, start_time, end_time, label, worksheet.title))

    if not entries:
        raise ValueError("The Adcom availability matrix did not contain any interviewer names.")
    notices.append(ImportNotice("warning", "adcom_roster_missing", "No separate Adcom roster is present; people with zero availability must be added in the People editor."))
    names = _unique_names(entries)
    workbook.close()
    return ParsedAvailability("adcom", entries, names, notices)


def reconcile_availability(
    availability: ParsedAvailability,
    slot_set: ParsedSlotSet,
    *,
    max_time_shift_minutes: int = 30,
) -> ParsedAvailability:
    """Map source availability to the authoritative slot list.

    Exact date/start matches are preferred. A unique nearest start on the same
    date may be reconciled within ``max_time_shift_minutes`` and is always
    reported as a warning.
    """

    exact = {(slot.start.date(), slot.start.time().replace(tzinfo=None)): slot for slot in slot_set.slots}
    by_date: dict[date, list[ParsedSlot]] = {}
    for slot in slot_set.slots:
        by_date.setdefault(slot.start.date(), []).append(slot)

    mapping: dict[tuple[date, time, time], ParsedSlot | None] = {}
    notices = list(availability.notices)
    reconciled_entries: list[ParsedAvailabilityEntry] = []

    for entry in availability.entries:
        source_key = (entry.source_date, entry.source_start, entry.source_end)
        if source_key not in mapping:
            mapped = exact.get((entry.source_date, entry.source_start))
            if mapped is None:
                candidates = []
                source_start_dt = datetime.combine(entry.source_date, entry.source_start)
                for candidate in by_date.get(entry.source_date, []):
                    delta = abs(int((candidate.start.replace(tzinfo=None) - source_start_dt).total_seconds() // 60))
                    if delta <= max_time_shift_minutes:
                        candidates.append((delta, candidate))
                candidates.sort(key=lambda item: (item[0], item[1].start))
                if candidates and (len(candidates) == 1 or candidates[0][0] < candidates[1][0]):
                    mapped = candidates[0][1]
                    notices.append(ImportNotice(
                        "warning",
                        "slot_time_reconciled",
                        f"Mapped {availability.group.title()} availability on {entry.source_date.isoformat()} from {_format_clock(entry.source_start)} to authoritative slot {_format_clock(mapped.start)}.",
                        entry.source_sheet,
                    ))
                else:
                    notices.append(ImportNotice("warning", "availability_slot_ignored", f"Ignored {availability.group.title()} availability for unmatched slot {entry.source_label!r}.", entry.source_sheet))
            mapping[source_key] = mapped

        mapped = mapping[source_key]
        if mapped is None:
            continue
        reconciled_entries.append(ParsedAvailabilityEntry(
            interviewer_name=entry.interviewer_name,
            source_date=entry.source_date,
            source_start=entry.source_start,
            source_end=entry.source_end,
            source_label=entry.source_label,
            source_sheet=entry.source_sheet,
            slot_id=mapped.id,
        ))

    # One person appearing multiple times in the same authoritative slot is a
    # source duplication, not additional availability.
    deduplicated: list[ParsedAvailabilityEntry] = []
    seen: set[tuple[str, str]] = set()
    duplicate_count = 0
    for entry in reconciled_entries:
        key = (entry.interviewer_name.casefold(), entry.slot_id or "")
        if key in seen:
            duplicate_count += 1
            continue
        seen.add(key)
        deduplicated.append(entry)
    if duplicate_count:
        notices.append(ImportNotice("warning", "duplicate_availability_removed", f"Removed {duplicate_count} duplicate interviewer/slot availability entries."))

    # Preserve the explicit roster even when a person has no matched slots.
    return ParsedAvailability(availability.group, deduplicated, list(availability.names), notices)
