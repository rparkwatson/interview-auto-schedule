"""Create and validate the administrative interview-period setup workbook."""

from __future__ import annotations

from datetime import datetime
from hashlib import sha256
from io import BytesIO
from pathlib import Path
import re
from typing import BinaryIO, Mapping, Sequence

from openpyxl import load_workbook
import xlsxwriter

from ..domain import Slot
from .source_parsers import ParsedSlotSet, WorkbookSource, parse_time_slot_workbook


PERIOD_TEMPLATE_SHEET = "Interview Periods"
PERIOD_TEMPLATE_METADATA_SHEET = "Template Metadata"
PERIOD_TEMPLATE_VERSION = "2"


class PeriodTemplateError(ValueError):
    """A completed period workbook is incomplete or belongs to another setup."""


def _format_interview_period(slot: Slot) -> str:
    day = slot.start.strftime("%a, %b %d").replace(" 0", " ")
    start_clock = slot.start.strftime("%I:%M %p").lstrip("0")
    end_clock = slot.end.strftime("%I:%M %p").lstrip("0")
    return f"{day} · {start_clock}–{end_clock}"


def period_fingerprint(
    slots: Sequence[Slot],
    *,
    timezone_name: str = "America/New_York",
) -> str:
    """Return a stable signature for the protected period structure."""

    rows = [
        "|".join(
            (
                str(slot.id),
                slot.start.isoformat(),
                slot.end.isoformat(),
                timezone_name,
            )
        )
        for slot in sorted(slots, key=lambda item: (item.start, item.id))
    ]
    return sha256("\n".join(rows).encode("utf-8")).hexdigest()


def period_template_filename(scenario: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", str(scenario).strip()).strip("_")
    return f"{(safe[:80] or 'scenario')}_interview_periods.xlsx"


def build_interview_period_template(
    slots: Sequence[Slot],
    *,
    scenario: str,
    student_available: Mapping[str, int] | None = None,
    adcom_available: Mapping[str, int] | None = None,
    timezone_name: str = "America/New_York",
) -> bytes:
    """Return a protected, fillable interview-period workbook in memory."""

    student_available = student_available or {}
    adcom_available = adcom_available or {}
    ordered_slots = sorted(slots, key=lambda item: (item.start, item.id))
    output = BytesIO()
    workbook = xlsxwriter.Workbook(
        output,
        {"in_memory": True, "remove_timezone": True},
    )
    workbook.set_properties(
        {
            "title": f"Interview period setup: {scenario}",
            "subject": "Enter the number of interviews possible in each period",
            "author": "Interview Scheduler v2",
            "comments": "Generated from the Student availability workbook.",
        }
    )
    navy = "#003B5C"
    pale_blue = "#DCE6F1"
    pale_yellow = "#FFF2CC"
    muted = "#F2F2F2"
    title_format = workbook.add_format(
        {
            "bold": True,
            "font_color": "#FFFFFF",
            "bg_color": navy,
            "font_size": 16,
            "align": "left",
            "valign": "vcenter",
        }
    )
    instruction_format = workbook.add_format(
        {"text_wrap": True, "font_color": navy, "bg_color": pale_blue}
    )
    header_format = workbook.add_format(
        {
            "bold": True,
            "font_color": "#FFFFFF",
            "bg_color": navy,
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "text_wrap": True,
        }
    )
    locked_format = workbook.add_format({"locked": True, "bg_color": muted})
    date_format = workbook.add_format(
        {"locked": True, "bg_color": muted, "num_format": "mmm d, yyyy"}
    )
    time_format = workbook.add_format(
        {"locked": True, "bg_color": muted, "num_format": "h:mm AM/PM"}
    )
    input_format = workbook.add_format(
        {
            "locked": False,
            "bg_color": pale_yellow,
            "border": 1,
            "align": "center",
            "num_format": "0",
        }
    )

    sheet = workbook.add_worksheet(PERIOD_TEMPLATE_SHEET)
    sheet.set_tab_color(navy)
    sheet.set_row(0, 24)
    sheet.merge_range(0, 0, 0, 7, f"Interview period setup · {scenario}", title_format)
    sheet.merge_range(
        1,
        0,
        1,
        7,
        (
            "Enter Interviews possible for every row. This is the number of interviews "
            "the scheduler will try to fill in that period. Enter 0 when a period will "
            "not be offered. Do not change dates or times."
        ),
        instruction_format,
    )
    sheet.set_row(1, 38)
    sheet.merge_range(
        2,
        0,
        2,
        7,
        "Yellow cells are the only cells that need input. All times are Eastern Time.",
        instruction_format,
    )

    headers = (
        "Period ID",
        "Date",
        "Start time",
        "End time",
        "Interview period",
        "Interviews possible",
        "Student available",
        "Adcom available",
    )
    header_row = 3
    first_data_row = header_row + 1
    for column, header in enumerate(headers):
        sheet.write(header_row, column, header, header_format)
    sheet.set_row(header_row, 32)

    for row_index, slot in enumerate(ordered_slots, start=first_data_row):
        sheet.write(row_index, 0, slot.id, locked_format)
        sheet.write_datetime(row_index, 1, slot.start.replace(tzinfo=None), date_format)
        sheet.write_datetime(row_index, 2, slot.start.replace(tzinfo=None), time_format)
        sheet.write_datetime(row_index, 3, slot.end.replace(tzinfo=None), time_format)
        sheet.write(
            row_index,
            4,
            _format_interview_period(slot),
            locked_format,
        )
        sheet.write_blank(row_index, 5, None, input_format)
        sheet.write(
            row_index,
            6,
            int(student_available.get(slot.id, 0)),
            locked_format,
        )
        sheet.write(
            row_index,
            7,
            int(adcom_available.get(slot.id, 0)),
            locked_format,
        )

    if ordered_slots:
        last_data_row = first_data_row + len(ordered_slots) - 1
        sheet.add_table(
            header_row,
            0,
            last_data_row,
            len(headers) - 1,
            {
                "name": "InterviewPeriodsTable",
                "style": "Table Style Medium 2",
                "columns": [{"header": header} for header in headers],
            },
        )
        sheet.data_validation(
            first_data_row,
            5,
            last_data_row,
            5,
            {
                "validate": "integer",
                "criteria": "between",
                "minimum": 0,
                "maximum": 999,
                "input_title": "Interviews possible",
                "input_message": "Enter a whole number. Use 0 to omit this period.",
                "error_title": "Enter a whole number",
                "error_message": "Use a whole number from 0 through 999.",
                "error_type": "stop",
            },
        )
    sheet.freeze_panes(first_data_row, 1)
    sheet.set_column(0, 0, 3, None, {"hidden": True})
    sheet.set_column(1, 1, 14)
    sheet.set_column(2, 3, 13)
    sheet.set_column(4, 4, 31)
    sheet.set_column(5, 5, 20)
    sheet.set_column(6, 7, 18)
    sheet.protect(
        "",
        {
            "select_locked_cells": False,
            "select_unlocked_cells": True,
            "autofilter": True,
            "sort": True,
        },
    )

    metadata = workbook.add_worksheet(PERIOD_TEMPLATE_METADATA_SHEET)
    metadata.write_column(
        0,
        0,
        (
            "template_type",
            "schema_version",
            "scenario",
            "timezone",
            "period_fingerprint",
            "period_count",
        ),
    )
    metadata.write_column(
        0,
        1,
        (
            "interview_period_setup",
            PERIOD_TEMPLATE_VERSION,
            scenario,
            timezone_name,
            period_fingerprint(ordered_slots, timezone_name=timezone_name),
            len(ordered_slots),
        ),
    )
    metadata.hide()
    workbook.close()
    return output.getvalue()


def _source_bytes(source: WorkbookSource) -> bytes:
    if isinstance(source, bytes):
        return source
    if isinstance(source, bytearray):
        return bytes(source)
    if isinstance(source, (str, Path)):
        return Path(source).read_bytes()
    stream: BinaryIO = source
    position = stream.tell() if hasattr(stream, "tell") else None
    payload = stream.read()
    if position is not None and hasattr(stream, "seek"):
        stream.seek(position)
    return payload


def parse_completed_interview_period_template(
    source: WorkbookSource,
    *,
    expected_slots: Sequence[Slot],
    year: int,
    timezone_name: str = "America/New_York",
) -> ParsedSlotSet:
    """Validate and parse counts from a generated interview-period workbook."""

    payload = _source_bytes(source)
    try:
        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    except Exception as exc:
        raise PeriodTemplateError("The completed file is not a readable Excel workbook.") from exc
    try:
        if PERIOD_TEMPLATE_METADATA_SHEET not in workbook.sheetnames:
            raise PeriodTemplateError(
                "This is not an interview-period template downloaded from this setup."
            )
        metadata_sheet = workbook[PERIOD_TEMPLATE_METADATA_SHEET]
        metadata = {
            str(key.value or "").strip(): str(value.value or "").strip()
            for key, value in metadata_sheet.iter_rows(min_col=1, max_col=2)
            if key.value is not None
        }
    finally:
        workbook.close()

    if metadata.get("template_type") != "interview_period_setup":
        raise PeriodTemplateError(
            "This is not an interview-period template downloaded from this setup."
        )
    if metadata.get("schema_version") != PERIOD_TEMPLATE_VERSION:
        raise PeriodTemplateError(
            "This template version is no longer supported. Download a new copy and try again."
        )
    expected_fingerprint = period_fingerprint(
        expected_slots,
        timezone_name=timezone_name,
    )
    if metadata.get("period_fingerprint") != expected_fingerprint:
        raise PeriodTemplateError(
            "This workbook belongs to a different availability setup. Download a new "
            "template from the current files and enter the counts there."
        )

    parsed = parse_time_slot_workbook(
        payload,
        year=year,
        timezone_name=timezone_name,
    )
    errors = [notice for notice in parsed.notices if notice.severity == "error"]
    if errors:
        details = "; ".join(notice.message for notice in errors[:5])
        raise PeriodTemplateError(
            "Complete every Interviews possible cell and correct any invalid counts. "
            + details
        )

    expected_by_id = {slot.id: slot for slot in expected_slots}
    parsed_by_id = {slot.id: slot for slot in parsed.slots}
    if set(parsed_by_id) != set(expected_by_id):
        raise PeriodTemplateError(
            "The period rows were added, removed, or renamed. Download a new template "
            "and change only the yellow count cells."
        )
    for slot_id, expected in expected_by_id.items():
        actual = parsed_by_id[slot_id]
        if actual.start != expected.start or actual.end != expected.end:
            raise PeriodTemplateError(
                "One or more dates or times were changed. Download a new template and "
                "change only the yellow count cells."
            )
    return parsed


__all__ = [
    "PERIOD_TEMPLATE_METADATA_SHEET",
    "PERIOD_TEMPLATE_SHEET",
    "PeriodTemplateError",
    "build_interview_period_template",
    "parse_completed_interview_period_template",
    "period_fingerprint",
    "period_template_filename",
]
