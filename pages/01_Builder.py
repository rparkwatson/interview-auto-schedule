import io
import re
import hashlib
import streamlit as st
import pandas as pd
import openpyxl
from adcom_contingency import maybe_convert_adcom_excel

from format_xlsx_core import parse_primary_and_adcom, build_formatted_workbook_bytes

# ────────────────────────────────────────────────────────────────────────────────
# Page setup
# ────────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Step 1 — Upload & Convert", layout="wide")
st.title("Step 1 — Upload → Convert → Edit → Generate")

st.subheader("Instructions")
st.markdown(
    """
    Use this app to process the AF/Regular and Adcom availability spreadsheets.
    Upload a .xlsx file for each group then click **Proceed**.
    
    After processing, you will be prompted to input date/time slot capacity and pre-assigned interview counts (if required). 

    Click **Generate formatted .xlsx** to be prompted to move to Step 2. 
    """
)
if st.session_state.get("proceeded"):
    st.divider()

# ────────────────────────────────────────────────────────────────────────────────
# Helpers
# ────────────────────────────────────────────────────────────────────────────────
def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip()).lower()


def _guess_name_column(df: pd.DataFrame) -> str | None:
    if df is None or df.empty:
        return None
    cols = [str(c) for c in df.columns]
    preferred = [
        "interviewer",
        "interviewer name",
        "name",
        "full name",
        "af name",
        "adcom",
        "adcom name",
    ]
    cols_norm = {_norm(c): c for c in cols}
    for p in preferred:
        if p in cols_norm:
            return cols_norm[p]
    for c in cols:
        cn = _norm(c)
        if any(tok in cn for tok in ["interviewer", "adcom", "name"]):
            return c
    for c in cols:
        if pd.api.types.is_object_dtype(df[c]):
            return c
    return cols[0] if cols else None


def _build_pre_table_from_names(names: pd.Series) -> pd.DataFrame:
    out = pd.DataFrame({"Interviewer": names.astype(str).str.strip().dropna().unique()})
    out = out[out["Interviewer"] != ""].copy()
    out.sort_values("Interviewer", inplace=True, key=lambda s: s.str.lower())
    out["Pre_Assigned_Count"] = 0
    out.reset_index(drop=True, inplace=True)
    return out


def _inject_preassigned_counts(
    xlsx_bytes: bytes,
    pre_df: pd.DataFrame,
    sheet_name: str | None,
    interviewer_header: str,
    count_header: str,
    header_row_index: int | None = None,
    scan_rows: int = 25,
    interviewer_aliases_extra: list[str] | None = None,
) -> bytes:
    """
    Auto-detect sheet and header row; match interviewer names case-insensitively.
    Adds count column if missing; writes counts.
    """
    bio = io.BytesIO(xlsx_bytes)
    wb = openpyxl.load_workbook(bio)

    def aliases(base_header: str, extras: list[str] | None) -> set[str]:
        base = {
            _norm(base_header),
            "interviewer",
            "interviewer name",
            "name",
            "interviewer_name",
            "interviewername",
            "adcom",
            "adcom name",
        }
        if extras:
            base |= {_norm(x) for x in extras}
        return base

    interviewer_aliases = aliases(interviewer_header, interviewer_aliases_extra)
    count_aliases = {
        _norm(count_header),
        "pre_assigned_count",
        "pre-assigned count",
        "pre assigned count",
    }

    candidate_ws = (
        [wb[sheet_name]] if sheet_name and sheet_name in wb.sheetnames else [wb[s] for s in wb.sheetnames]
    )

    found = None  # (ws, hdr_row_idx, interviewer_col, count_col)
    for ws in candidate_ws:
        max_scan = min(scan_rows, ws.max_row if ws.max_row else 1)
        for r in range(1, max_scan + 1):
            headers = {}
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    headers[_norm(val)] = c
            if not headers:
                continue
            interviewer_col = None
            for k, colidx in headers.items():
                if k in interviewer_aliases:
                    interviewer_col = colidx
                    break
            if interviewer_col is None:
                continue
            count_col = None
            for k, colidx in headers.items():
                if k in count_aliases:
                    count_col = colidx
                    break
            found = (ws, r, interviewer_col, count_col)
            break
        if found:
            break

    if not found:
        raise ValueError(
            f"Couldn’t locate an interviewer column in the first {scan_rows} row(s) of any sheet. "
            f"Tried aliases: {sorted(interviewer_aliases)}. Sheets: {', '.join(wb.sheetnames)}"
        )

    ws, hdr_row_idx, interviewer_col, count_col = found

    # Respect explicit header row if it still contains interviewer header
    if header_row_index:
        headers_on_provided = {_norm(c.value): c.column for c in ws[header_row_index] if c.value}
        for k in interviewer_aliases:
            if k in headers_on_provided:
                hdr_row_idx = header_row_index
                interviewer_col = headers_on_provided[k]
                count_col = None
                for k2 in count_aliases:
                    if k2 in headers_on_provided:
                        count_col = headers_on_provided[k2]
                        break
                break

    if count_col is None:
        count_col = ws.max_column + 1
        ws.cell(row=hdr_row_idx, column=count_col, value=count_header)

    # Build lookup
    pre_df = pre_df.copy()
    pre_df["Interviewer"] = pre_df["Interviewer"].astype(str).str.strip()
    pre_df["Pre_Assigned_Count"] = (
        pd.to_numeric(pre_df["Pre_Assigned_Count"], errors="coerce").fillna(0).astype(int)
    )
    lookup = {_norm(n): int(v) for n, v in zip(pre_df["Interviewer"], pre_df["Pre_Assigned_Count"])}

    applied, unmatched = 0, set(lookup.keys())
    for r in range(hdr_row_idx + 1, ws.max_row + 1):
        raw_name = ws.cell(row=r, column=interviewer_col).value
        if raw_name is None:
            continue
        key = _norm(raw_name)
        if key in lookup:
            ws.cell(row=r, column=count_col, value=lookup[key])
            applied += 1
            unmatched.discard(key)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    if "pre_stats" not in st.session_state:
        st.session_state.pre_stats = []
    st.session_state.pre_stats.append(
        {"sheet": ws.title, "header_row": hdr_row_idx, "applied": applied, "unmatched": sorted(unmatched)}
    )
    return out.read()


def _clear_everything():
    for k in [
        "primary",
        "adcom",
        "primary_bytes",
        "adcom_bytes",  # NEW cached bytes
        "master_df",
        "max_df",
        "adcom_df",
        "reg_pre_df",
        "adcom_pre_df",
        "reg_name_col",
        "adcom_name_col",
        "formatted_xlsx",
        "pre_stats",
        "last_gen_fp",  # NEW stored fingerprint
        "proceeded",
        # any widget keys to reset explicitly:
        "reg_fill_val",
        "adcom_fill_val",
        "max_editor",
        "reg_pre_editor",
        "adcom_pre_editor",
        "reg_name_col_select",
        "adcom_name_col_select",
        # recode extras
        "recode_names_norm",
        "recode_stats",
    ]:
        st.session_state.pop(k, None)


# ────────────────────────────────────────────────────────────────────────────────
# Re-code Regular → Adcom helpers
# ────────────────────────────────────────────────────────────────────────────────

def _recode_names_normalized(names: list[str]) -> list[str]:
    return [_norm(n) for n in names if str(n).strip()]


def _recode_pre_tables(
    reg_df: pd.DataFrame, adcom_df: pd.DataFrame, names_norm: list[str]
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Move any Interviewer in names_norm from Regular pre table to Adcom pre table."""
    if reg_df is None or reg_df.empty or not names_norm:
        return reg_df, adcom_df

    reg_df = reg_df.copy()
    adcom_df = adcom_df.copy()

    # Build fast lookup on normalized interviewer names
    reg_df["_k"] = (
        reg_df["Interviewer"].astype(str).str.strip().str.lower().str.replace(r"\s+", " ", regex=True)
    )
    adcom_df["_k"] = (
        adcom_df["Interviewer"].astype(str).str.strip().str.lower().str.replace(r"\s+", " ", regex=True)
    )

    to_move = reg_df[reg_df["_k"].isin(names_norm)].copy()
    if to_move.empty:
        reg_df.drop(columns=["_k"], inplace=True)
        adcom_df.drop(columns=["_k"], inplace=True)
        return reg_df, adcom_df

    # Merge/move rows: if already present on Adcom, keep the larger count
    to_move = to_move.drop(columns=["_k"])  # strip helper col
    adcom_df = adcom_df.drop(columns=["_k"])  # strip helper col

    if "Pre_Assigned_Count" in to_move.columns and "Pre_Assigned_Count" in adcom_df.columns:
        adcom_df = (
            pd.concat([adcom_df, to_move], ignore_index=True)
            .sort_values("Interviewer")
            .groupby("Interviewer", as_index=False)["Pre_Assigned_Count"]
            .max()
        )
    else:
        adcom_df = pd.concat([adcom_df, to_move], ignore_index=True)

    # Remove from Regular
    reg_df = reg_df[
        ~reg_df["Interviewer"].astype(str)
        .str.strip()
        .str.lower()
        .str.replace(r"\s+", " ", regex=True)
        .isin(names_norm)
    ].copy()

    return reg_df, adcom_df


def _recode_regular_to_adcom_in_workbook(
    xlsx_bytes: bytes,
    *,
    reg_sheet: str,
    adcom_sheet: str,
    reg_name_header: str,
    adcom_name_header: str,
    header_row_index: int,
) -> bytes:
    """
    Open the built workbook and move any rows whose interviewer (on Regular) matches
    the configured names into the Adcom sheet. Column mapping is done by header name.
    """
    names_norm: list[str] = st.session_state.get("recode_names_norm", [])
    if not names_norm:
        return xlsx_bytes

    bio = io.BytesIO(xlsx_bytes)
    wb = openpyxl.load_workbook(bio)

    if reg_sheet not in wb.sheetnames or adcom_sheet not in wb.sheetnames:
        st.warning(
            f"Re-code skipped: sheet(s) not found (Regular='{reg_sheet}', Adcom='{adcom_sheet}')."
        )
        return xlsx_bytes

    ws_reg = wb[reg_sheet]
    ws_adc = wb[adcom_sheet]

    # Build header maps for both sheets (normalized header -> column index)
    def header_map(ws, hdr_row):
        m = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=hdr_row, column=c).value
            if v is not None:
                m[_norm(v)] = c
        return m

    hdr_reg = header_map(ws_reg, header_row_index)
    hdr_adc = header_map(ws_adc, header_row_index)

    # Find interviewer columns (with aliases)
    def find_col(hdr_map: dict, primary_header: str) -> int | None:
        aliases = {
            _norm(primary_header),
            "interviewer", "interviewer name", "interviewer_name", "interviewername",
            "name",
            # common Regular/AF variants
            "af", "af name", "af_name", "af interviewer", "af_interviewer", "af interviewer name",
            "regular", "regular name", "regular interviewer",
            # sometimes sheets label the column as 'adcom' even on regular tab
            "adcom", "adcom name",
        }
        # exact match first
        for a in aliases:
            if a in hdr_map:
                return hdr_map[a]
        # fuzzy fallback: any header containing 'interview'
        for k in hdr_map:
            if "interview" in k:
                return hdr_map[k]
        # last-resort: a plain 'name' column if present
        if "name" in hdr_map:
            return hdr_map["name"]
        return None

    col_reg_name = find_col(hdr_reg, reg_name_header)
    col_adc_name = find_col(hdr_adc, adcom_name_header)  # not strictly needed but nice to have

    if not col_reg_name:
        # Try to auto-detect the header row within the first N rows
        SCAN_ROWS = 25
        max_scan_reg = min(SCAN_ROWS, ws_reg.max_row if ws_reg.max_row else 1)
        for r in range(1, max_scan_reg + 1):
            hdr_try = header_map(ws_reg, r)
            col_try = find_col(hdr_try, reg_name_header)
            if col_try:
                hdr_reg = hdr_try
                col_reg_name = col_try
                header_row_index = r
                break

        # If still not found, heuristic: detect the name column by looking for the target names in cells
        if not col_reg_name:
            names_norm_set = set(names_norm)
            max_rows_check = min(ws_reg.max_row or 1, (header_row_index or 1) + 200)
            for c in range(1, ws_reg.max_column + 1):
                hit = False
                for r in range((header_row_index or 1) + 1, max_rows_check + 1):
                    val = ws_reg.cell(row=r, column=c).value
                    if val is None:
                        continue
                    if _norm(val) in names_norm_set:
                        col_reg_name = c
                        hit = True
                        break
                if hit:
                    break

        # Also try to auto-detect the Adcom header row if needed (for column mapping)
        col_adc_name = find_col(hdr_adc, adcom_name_header)
        if not col_adc_name:
            max_scan_adc = min(SCAN_ROWS, ws_adc.max_row if ws_adc.max_row else 1)
            for r in range(1, max_scan_adc + 1):
                hdr_try = header_map(ws_adc, r)
                col_try = find_col(hdr_try, adcom_name_header)
                if col_try:
                    hdr_adc = hdr_try
                    col_adc_name = col_try
                    break

        if not col_reg_name:
            st.warning(
                f"Re-code skipped: couldn't find interviewer column on Regular sheet. "
                f"Checked rows 1–{max_scan_reg}. Headers near provided row {header_row_index}: {sorted(hdr_reg.keys())[:12]}"
            )
            return xlsx_bytes

    # Collect rows to move
    rows_to_move = []
    for r in range(header_row_index + 1, ws_reg.max_row + 1):
        val = ws_reg.cell(row=r, column=col_reg_name).value
        if val is None:
            continue
        if _norm(val) in names_norm:
            rows_to_move.append(r)

    if not rows_to_move:
        st.info("Re-code: no matching Regular interviewers found to move.")
        return xlsx_bytes

    shared_keys = set(hdr_reg.keys()) & set(hdr_adc.keys())

    moved_names = set()
    for r in rows_to_move:
        # Append a new row to Adcom and map by header names
        dst_row = ws_adc.max_row + 1
        for key in shared_keys:
            src_c = hdr_reg[key]
            dst_c = hdr_adc[key]
            ws_adc.cell(row=dst_row, column=dst_c, value=ws_reg.cell(row=r, column=src_c).value)
        if col_adc_name:
            moved_names.add(_norm(ws_reg.cell(row=r, column=col_reg_name).value))

    # Delete moved rows from Regular (bottom-up so indices stay valid)
    for r in sorted(rows_to_move, reverse=True):
        ws_reg.delete_rows(r, 1)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    # Surface stats
    if "recode_stats" not in st.session_state:
        st.session_state.recode_stats = []
    st.session_state.recode_stats.append(
        {
            "moved_count": len(rows_to_move),
            "moved_names": sorted(moved_names),
            "from_sheet": reg_sheet,
            "to_sheet": adcom_sheet,
        }
    )
    return out.read()


# ────────────────────────────────────────────────────────────────────────────────
# Fingerprint helpers (to gate download / step 2)
# ────────────────────────────────────────────────────────────────────────────────

def _df_to_stable_str(df: pd.DataFrame | None, name: str) -> str:
    if df is None or (hasattr(df, "empty") and df.empty):
        return f"{name}:<empty>"
    # Stable: include columns + CSV values (no index)
    return f"{name}:{','.join(map(str, df.columns))}\n{df.to_csv(index=False)}"


def _fingerprint_generation_inputs() -> str:
    """Hash only the inputs that affect the formatted workbook."""
    parts: list[str] = []
    parts.append(_df_to_stable_str(st.session_state.get("max_df"), "max_df"))
    parts.append(_df_to_stable_str(st.session_state.get("reg_pre_df"), "reg_pre_df"))
    parts.append(_df_to_stable_str(st.session_state.get("adcom_pre_df"), "adcom_pre_df"))
    return hashlib.sha256("||".join(parts).encode("utf-8")).hexdigest()


def _invalidate_if_inputs_changed(show_message: bool = True):
    """If a workbook exists but inputs changed, clear it and optionally inform the user."""
    current_fp = _fingerprint_generation_inputs()
    prev_fp = st.session_state.get("last_gen_fp")
    if st.session_state.get("formatted_xlsx") and prev_fp and current_fp != prev_fp:
        st.session_state.formatted_xlsx = None
        if show_message:
            st.info(
                "Inputs changed (Max_Pairs or Pre_Assigned_Counts). Please click **Generate Workbook for Step 2** again."
            )


# ────────────────────────────────────────────────────────────────────────────────
# Sidebar
# ────────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    with st.expander("Settings", expanded=False):
        st.markdown("### Header Row Adjustment")
        header_row = st.number_input("Header row (0-based)", min_value=0, max_value=200, value=3, step=1)

    with st.expander("Pre Assigned Data Match Fields", expanded=False):
        st.markdown("### Inject targets (pre-set)")
        reg_sheet = st.text_input("Regular sheet name", value="Master_Availability_Sheet")
        reg_name_header = st.text_input("Regular interviewer header", value="Interviewer_Name")
        reg_count_header = st.text_input("Regular count header", value="Pre_Assigned_Count")

        adcom_sheet = st.text_input("Adcom sheet name", value="Adcom_Availability")
        adcom_name_header = st.text_input("Adcom interviewer header", value="Interviewer_Name")
        adcom_count_header = st.text_input("Adcom count header", value="Pre_Assigned_Count")

    # NEW: Re-code UI
    with st.expander("Re-code Regular ➜ Adcom (by name)", expanded=False):
        st.markdown(
            "If present on the Regular sheet, these interviewers will be moved to the Adcom sheet (before counts are injected)."
        )
        recode_name_1 = st.text_input("Name #1", value="Kimberly Oliva")
        recode_name_2 = st.text_input("Name #2", value="Priya Trauber")
        recode_extra_csv = st.text_input("Additional names (comma-separated, optional)", value="")
        recode_list = [recode_name_1, recode_name_2] + [x.strip() for x in recode_extra_csv.split(",") if x.strip()]
        st.session_state.recode_names_norm = _recode_names_normalized(recode_list)

    if st.button("Clear uploads", key="clear_uploads_btn", type="primary"):
        _clear_everything()
        st.rerun()

# ────────────────────────────────────────────────────────────────────────────────
# Uploads
# ────────────────────────────────────────────────────────────────────────────────
col1, col2 = st.columns(2)
with col1:
    primary = st.file_uploader("**AF Availability** Excel", type=["xlsx", "xls"], key="primary")
with col2:
    adcom = st.file_uploader("**Adcom Availability** Excel", type=["xlsx", "xls"], key="adcom")

# ────────────────────────────────────────────────────────────────────────────────
# Parse 
# ────────────────────────────────────────────────────────────────────────────────
uploads_present = (st.session_state.get("primary") is not None) and (st.session_state.get("adcom") is not None)
proceeded = bool(st.session_state.get("proceeded", False))

show_proceed = (not proceeded) or (not uploads_present)

proceed_clicked = False
if show_proceed:
    proceed_clicked = st.button(
        "Proceed",
        key="proceed_btn",
        type="primary",
        disabled=not uploads_present,
        help=None if uploads_present else "Upload both files to enable Proceed",
    )
else:
    st.caption("✅ Already processed. Clear or change an upload to re-enable **Proceed**.")

if proceed_clicked and uploads_present:
    try:
        # Read uploads once and cache bytes for reuse (fix EOF / engine sniff error)
        p_up = st.session_state["primary"]
        a_up = st.session_state["adcom"]
        p_bytes = p_up.getvalue() if hasattr(p_up, "getvalue") else p_up.read()
        a_bytes = a_up.getvalue() if hasattr(a_up, "getvalue") else a_up.read()

        # ── MINIMAL CONTINGENCY: auto-convert Adcom if it is the matrix format ──
        try:
            # If you have an example workbook, pass it as io.BytesIO(example_bytes) instead of None
            a_bytes_conv = maybe_convert_adcom_excel(
                file_like=io.BytesIO(a_bytes),
                example_file_like=None,  # ← optional: io.BytesIO(example_bytes)
                default_year=2025,
            )
        except Exception as conv_err:
            st.warning(f"Adcom contingency conversion skipped: {conv_err}")
            a_bytes_conv = a_bytes
        # ────────────────────────────────────────────────────────────────────────

        # Persist the original AF bytes and the (possibly converted) Adcom bytes
        st.session_state.primary_bytes = p_bytes
        st.session_state.adcom_bytes = a_bytes_conv

        # Fresh BytesIO for parsing
        p_bio = io.BytesIO(p_bytes)
        a_bio = io.BytesIO(a_bytes_conv)

        master_df, max_df, adcom_df = parse_primary_and_adcom(p_bio, a_bio, header_row)

        if "Max_Pairs" not in max_df.columns:
            max_df["Max_Pairs"] = 0
        max_df["Max_Pairs"] = max_df["Max_Pairs"].fillna(0).astype(int)

        # Persist
        st.session_state.master_df = master_df
        st.session_state.max_df = max_df
        st.session_state.adcom_df = adcom_df

        # Initialize pre-assign tables
        st.session_state.reg_name_col = _guess_name_column(master_df)
        st.session_state.adcom_name_col = _guess_name_column(adcom_df)

        reg_names = (
            master_df[st.session_state.reg_name_col]
            if st.session_state.reg_name_col in master_df.columns
            else pd.Series([], dtype=str)
        )
        adcom_names = (
            adcom_df[st.session_state.adcom_name_col]
            if st.session_state.adcom_name_col in adcom_df.columns
            else pd.Series([], dtype=str)
        )
        st.session_state.reg_pre_df = _build_pre_table_from_names(reg_names)
        st.session_state.adcom_pre_df = _build_pre_table_from_names(adcom_names)

        # NEW: Apply Regular ➜ Adcom re-code to the on-screen pre-assignment tables so the UI reflects the move
        if st.session_state.get("recode_names_norm"):
            new_reg_pre, new_adc_pre = _recode_pre_tables(
                st.session_state.reg_pre_df,
                st.session_state.adcom_pre_df,
                st.session_state.recode_names_norm,
            )
            st.session_state.reg_pre_df = new_reg_pre
            st.session_state.adcom_pre_df = new_adc_pre

        st.session_state.proceeded = True
        st.success("Parsed successfully. Edit Max_Pairs and Pre-Assigned tables below.")

        st.rerun()

    except Exception as e:
        st.session_state.proceeded = False
        st.error(f"Parse failed: {e}")


# ────────────────────────────────────────────────────────────────────────────────
# Summary dashboard (after Proceed)
# ────────────────────────────────────────────────────────────────────────────────

def _interview_stats(df: pd.DataFrame, label: str, name_col: str | None) -> dict:
    """Compute basic stats; 'interviews' = non-empty rows in the interviewer column."""
    if df is None or df.empty:
        return {"File": label, "Rows": 0, "Interviews": 0, "Unique Interviewers": 0}

    col = name_col if name_col in df.columns else _guess_name_column(df)
    names = df[col].astype(str).str.strip().replace({"nan": ""})
    nonempty = names[names != ""]
    return {
        "File": label,
        "Rows": int(len(df)),
        "Interviews": int(len(nonempty)),
        "Unique Interviewers": int(nonempty.nunique()),
    }


if "master_df" in st.session_state and "adcom_df" in st.session_state:
    with st.container(border=True):
        st.subheader("Summary — Uploaded Files")

        af_stats = _interview_stats(
            st.session_state.master_df,
            "AF Availability",
            st.session_state.get("reg_name_col"),
        )
        adcom_stats = _interview_stats(
            st.session_state.adcom_df,
            "Adcom Availability",
            st.session_state.get("adcom_name_col"),
        )

        c1, c2 = st.columns(2)
        with c1:
            st.metric("AF Interviews", af_stats["Interviews"])
            st.metric("AF Unique Interviewers", af_stats["Unique Interviewers"])
        with c2:
            st.metric("Adcom Interviews", adcom_stats["Interviews"])
            st.metric("Adcom Unique Interviewers", adcom_stats["Unique Interviewers"])

if st.session_state.get("proceeded"):
    st.divider()

# ────────────────────────────────────────────────────────────────────────────────
# Max pairs editor
# ────────────────────────────────────────────────────────────────────────────────
if "max_df" in st.session_state:
    st.subheader("Date/Time Slot Capacity")
    left, right = st.columns([1, 3])
    with left:
        global_default = st.number_input("Apply global default", min_value=0, max_value=999, value=0)
        if st.button("Apply to all"):
            st.session_state.max_df["Max_Pairs"] = int(global_default)

    edited = st.data_editor(
        st.session_state.max_df,
        hide_index=True,
        num_rows="fixed",
        column_config={
            "Date_Time": st.column_config.TextColumn(disabled=True),
            "Max_Pairs": st.column_config.NumberColumn(min_value=0, max_value=999, step=1),
        },
        key="max_editor",
        width='stretch',
    )
    st.session_state.max_df = edited

if st.session_state.get("proceeded"):
    st.divider()

# ────────────────────────────────────────────────────────────────────────────────
# Pre-Assigned Counts (paste-friendly tables)
# ────────────────────────────────────────────────────────────────────────────────
if "master_df" in st.session_state and "adcom_df" in st.session_state:
    st.subheader("Pre-Assigned Counts")

    c1, c2 = st.columns(2)
    with c1:
        reg_name_col = st.selectbox(
            "Regular: pick the interviewer name column",
            options=list(st.session_state.master_df.columns),
            index=(
                max(0, list(st.session_state.master_df.columns).index(st.session_state.reg_name_col))
                if st.session_state.reg_name_col in st.session_state.master_df.columns
                else 0
            ),
            key="reg_name_col_select",
        )
    with c2:
        adcom_name_col = st.selectbox(
            "Adcom: pick the interviewer name column",
            options=list(st.session_state.adcom_df.columns),
            index=(
                max(0, list(st.session_state.adcom_df.columns).index(st.session_state.adcom_name_col))
                if st.session_state.adcom_name_col in st.session_state.adcom_df.columns
                else 0
            ),
            key="adcom_name_col_select",
        )

    # Rebuild tables if selection changed
    if reg_name_col != st.session_state.reg_name_col:
        st.session_state.reg_name_col = reg_name_col
        st.session_state.reg_pre_df = _build_pre_table_from_names(
            st.session_state.master_df[reg_name_col]
        )
    if adcom_name_col != st.session_state.adcom_name_col:
        st.session_state.adcom_name_col = adcom_name_col
        st.session_state.adcom_pre_df = _build_pre_table_from_names(
            st.session_state.adcom_df[adcom_name_col]
        )

    st.caption("Tip: paste an entire column of numbers into **Pre_Assigned_Count** from Excel/Sheets.")

    colA, colB = st.columns(2)
    with colA:
        st.markdown("**Regular**")
        leftA, rightA = st.columns([1, 3])
        with leftA:
            reg_fill = st.number_input("Set all (Regular)", min_value=0, max_value=999, value=0, key="reg_fill_val")
            if st.button("Apply to all (Regular)"):
                st.session_state.reg_pre_df["Pre_Assigned_Count"] = int(reg_fill)
        st.session_state.reg_pre_df = st.data_editor(
            st.session_state.reg_pre_df,
            hide_index=True,
            num_rows="dynamic",
            column_config={
                "Interviewer": st.column_config.TextColumn(disabled=True),
                "Pre_Assigned_Count": st.column_config.NumberColumn(min_value=0, step=1),
            },
            key="reg_pre_editor",
            width='stretch',
        )

    with colB:
        st.markdown("**Adcom**")
        leftB, rightB = st.columns([1, 3])
        with leftB:
            adcom_fill = st.number_input("Set all (Adcom)", min_value=0, max_value=999, value=0, key="adcom_fill_val")
            if st.button("Apply to all (Adcom)"):
                st.session_state.adcom_pre_df["Pre_Assigned_Count"] = int(adcom_fill)
        st.session_state.adcom_pre_df = st.data_editor(
            st.session_state.adcom_pre_df,
            hide_index=True,
            num_rows="dynamic",
            column_config={
                "Interviewer": st.column_config.TextColumn(disabled=True),
                "Pre_Assigned_Count": st.column_config.NumberColumn(min_value=0, step=1),
            },
            key="adcom_pre_editor",
            width='stretch',
        )

# ────────────────────────────────────────────────────────────────────────────────
# Invalidate previously generated workbook if inputs changed (after editors)
# ────────────────────────────────────────────────────────────────────────────────
if st.session_state.get("proceeded"):
    _invalidate_if_inputs_changed(show_message=True)

# ────────────────────────────────────────────────────────────────────────────────
# Generate workbook (+ inject both tables if present)
# ────────────────────────────────────────────────────────────────────────────────
if "max_df" in st.session_state and st.button("Generate Workbook for Step 2", type="primary"):
    try:
        # Guard: ensure we have original uploaded bytes
        if not st.session_state.get("primary_bytes") or not st.session_state.get("adcom_bytes"):
            raise RuntimeError("Original uploads not available. Please re-upload and click Proceed again.")

        # Always use fresh streams at position 0
        primary_src = io.BytesIO(st.session_state["primary_bytes"])
        adcom_src = io.BytesIO(st.session_state["adcom_bytes"])

        overrides = dict(zip(st.session_state.max_df["Date_Time"], st.session_state.max_df["Max_Pairs"]))

        xlsx_bytes, _parts = build_formatted_workbook_bytes(
            primary_file_like=primary_src,
            adcom_file_like=adcom_src,
            header_row=header_row,
            max_pairs_overrides=overrides,
        )

        header_row_index = header_row + 1  # Excel is 1-based
        st.session_state.pre_stats = []  # reset
        st.session_state.recode_stats = []  # reset

        # NEW: Re-code Regular ➜ Adcom rows in the built workbook so count injection lands on the right sheet
        if st.session_state.get("recode_names_norm"):
            xlsx_bytes = _recode_regular_to_adcom_in_workbook(
                xlsx_bytes=xlsx_bytes,
                reg_sheet=reg_sheet.strip() or "Master_Availability_Sheet",
                adcom_sheet=adcom_sheet.strip() or "Adcom_Availability",
                reg_name_header=reg_name_header.strip() or "Interviewer",
                adcom_name_header=adcom_name_header.strip() or "Interviewer",
                header_row_index=header_row_index,
            )

        # Inject Regular
        if "reg_pre_df" in st.session_state and not st.session_state.reg_pre_df.empty:
            xlsx_bytes = _inject_preassigned_counts(
                xlsx_bytes=xlsx_bytes,
                pre_df=st.session_state.reg_pre_df,
                sheet_name=reg_sheet.strip() or None,
                interviewer_header=reg_name_header.strip() or "Interviewer",
                count_header=reg_count_header.strip() or "Pre_Assigned_Count",
                header_row_index=header_row_index,
                interviewer_aliases_extra=["regular", "af", "af interviewer", "af name"],
            )

        # Inject Adcom
        if "adcom_pre_df" in st.session_state and not st.session_state.adcom_pre_df.empty:
            xlsx_bytes = _inject_preassigned_counts(
                xlsx_bytes=xlsx_bytes,
                pre_df=st.session_state.adcom_pre_df,
                sheet_name=adcom_sheet.strip() or None,
                interviewer_header=adcom_name_header.strip() or "Interviewer",
                count_header=adcom_count_header.strip() or "Pre_Assigned_Count",
                header_row_index=header_row_index,
                interviewer_aliases_extra=["adcom", "adcom interviewer", "adcom name"],
            )

        st.session_state.formatted_xlsx = xlsx_bytes
        st.session_state.last_gen_fp = _fingerprint_generation_inputs()  # record fingerprint used

        # Surface stats
        if getattr(st.session_state, "pre_stats", None):
            for s in st.session_state.pre_stats:
                msg = f"Applied {s['applied']} on sheet '{s['sheet']}', header row {s['header_row']}."
                if s["unmatched"]:
                    msg += f" Unmatched: {', '.join(s['unmatched'][:8])}"
                    if len(s["unmatched"]) > 8:
                        msg += f" (+{len(s['unmatched']) - 8} more)"
                st.info(msg)

        # Surface re-code stats
        if getattr(st.session_state, "recode_stats", None):
            for s in st.session_state.recode_stats:
                moved_preview = ", ".join(s["moved_names"][:6]) if s["moved_names"] else "—"
                extra = "" if len(s["moved_names"]) <= 6 else f" (+{len(s['moved_names']) - 6} more)"
                st.info(
                    f"Re-coded {s['moved_count']} row(s) from '{s['from_sheet']}' to '{s['to_sheet']}'. Names: {moved_preview}{extra}"
                )

        st.success("Workbook generated!")

    except Exception as e:
        st.error(f"Generation failed: {e}")

# ────────────────────────────────────────────────────────────────────────────────
# Download & Next step (gated by fingerprint validity)
# ────────────────────────────────────────────────────────────────────────────────
has_build = st.session_state.get("formatted_xlsx") is not None

if has_build:
    current_fp = _fingerprint_generation_inputs()
    valid_build = st.session_state.get("last_gen_fp") == current_fp

    # Message only when a file exists; choose success vs warning based on validity
    if valid_build:
        st.success("**Next step**\n\nYour file is ready. Continue to scheduling:")
    else:
        st.warning(
            "Edits detected after the last build. Click **Generate Workbook for Step 2** to rebuild."
        )

    # Download button (disabled if invalid)
    st.download_button(
        "Download Formatted Excel",
        data=st.session_state["formatted_xlsx"],
        file_name="formatted.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        disabled=not valid_build,
        help=None
        if valid_build
        else "Edits detected. Click “Generate Workbook for Step 2” to rebuild.",
    )

    # Step 2 button (disabled if invalid)
    if valid_build:
        if st.button("➡️ Open Step 2 — Scheduler", type="secondary", width='stretch'):
            st.switch_page("pages/02_Scheduler.py")
    else:
        st.button(
            "➡️ Open Step 2 — Scheduler",
            type="secondary",
            width='stretch',
            disabled=True,
            help="Edits detected. Click “Generate Workbook for Step 2” to rebuild.",
        )
