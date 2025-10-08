# adcom_contingency.py
import io
import re
from collections import defaultdict, OrderedDict
from datetime import datetime
from typing import Dict, List, Tuple, Optional

import pandas as pd
from openpyxl import load_workbook


# ────────────────────────────────────────────────────────────────────────────────
# Detection logic
# ────────────────────────────────────────────────────────────────────────────────
DAY_NAMES = r"(Mon|Tue|Tues|Wed|Thu|Thur|Fri|Sat|Sun|Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)"
HEADER_RGX = re.compile(
    rf"{DAY_NAMES}\s*\(\s*\d{{1,2}}/\d{{1,2}}\s*\)\s*[\r\n]+.*?\d{{1,2}}:\d{{2}}\s*[-–]\s*\d{{1,2}}:\d{{2}}\s*(?:[ap]\.?m\.?)?",
    re.IGNORECASE | re.DOTALL,
)
TIME_RANGE_RGX = re.compile(
    r"^\s*(\d{1,2}):(\d{2})\s*[-–]\s*(\d{1,2}):(\d{2})\s*([ap]\.?m\.?)?\s*$",
    re.IGNORECASE,
)
DATE_RGX = re.compile(r"^\s*([A-Za-z]+)\s*\(\s*(\d{1,2})/(\d{1,2})\s*\)\s*$")
DOW_ABBR = {
    "monday": "Mon", "tuesday": "Tue", "wednesday": "Wed", "thursday": "Thu",
    "friday": "Fri", "saturday": "Sat", "sunday": "Sun",
    "mon": "Mon", "tue": "Tue", "tues": "Tue", "wed": "Wed",
    "thu": "Thu", "thur": "Thu", "fri": "Fri", "sat": "Sat", "sun": "Sun",
}


def _normalize_ampm(s: str) -> str:
    s = (s or "").lower().replace(".", "")
    if s.endswith("am") or s.endswith("pm"):
        return s[-2:].upper()
    return ""


def _is_matrix_header_cell(val: str) -> bool:
    val = str(val or "").strip()
    return bool(val and HEADER_RGX.search(val))


def _detect_matrix_sheet(df: pd.DataFrame) -> Tuple[bool, List[int], Dict[int, List[Tuple[int, str]]]]:
    """
    Return (is_matrix_like, header_row_idxs, header_cells_by_row)
    header_cells_by_row[row] -> list[(col_idx, header_text)]
    """
    header_rows, header_cells = [], {}
    for i in range(df.shape[0]):
        matches = []
        for j in range(df.shape[1]):
            if _is_matrix_header_cell(df.iat[i, j]):
                matches.append((j, str(df.iat[i, j])))
        # section header row usually contains many header-like cells (columns of time/date)
        if len(matches) >= 3:
            header_rows.append(i)
            header_cells[i] = matches
    return (len(header_rows) > 0), header_rows, header_cells


# ────────────────────────────────────────────────────────────────────────────────
# Parsing helpers
# ────────────────────────────────────────────────────────────────────────────────
def _parse_header_cell(text: str, year: int) -> Tuple[str, str, str, str]:
    """
    Returns (date_key, display_date_header, start_label, end_label)
      - date_key: "Mon 10/27/2025"
      - display_date_header: "Monday (10/27)"
      - start_label: "8:00 AM"
      - end_label: "9:30 AM"
    """
    parts = [p.strip() for p in str(text).splitlines() if p and str(p).strip()]
    if len(parts) != 2:
        raise ValueError("Header cell not in expected 2-line format")

    m_date = DATE_RGX.match(parts[0])
    m_time = TIME_RANGE_RGX.match(parts[1])
    if not (m_date and m_time):
        raise ValueError("Header cell didn't match date/time patterns")

    dow_raw, mm, dd = m_date.groups()
    h1, m1, h2, m2, ampm = m_time.groups()
    ampm_norm = _normalize_ampm(ampm)

    start_label = f"{int(h1)}:{m1}"
    end_label = f"{int(h2)}:{m2}"
    if ampm_norm:
        start_label = f"{start_label} {ampm_norm}"
        end_label = f"{end_label} {ampm_norm}"

    dow_abbr = DOW_ABBR.get(dow_raw.lower(), dow_raw[:3].title())
    date_key = f"{dow_abbr} {int(mm)}/{int(dd)}/{year}"
    display_date_header = f"{dow_raw.title()} ({int(mm)}/{int(dd)})"
    return date_key, display_date_header, start_label, end_label


def _build_timesheets(
    df: pd.DataFrame,
    header_rows: List[int],
    header_cells: Dict[int, List[Tuple[int, str]]],
    year: int,
):
    """
    timesheets[time_label][date_key] = [names...]
    display_headers_by_date[date_key] = "Monday (10/27)"
    """
    timesheets: Dict[str, "OrderedDict[str, List[str]]"] = defaultdict(OrderedDict)
    display_headers_by_date: Dict[str, str] = {}

    for idx, hr in enumerate(header_rows):
        next_hr = header_rows[idx + 1] if idx + 1 < len(header_rows) else df.shape[0]
        name_rows = range(hr + 1, next_hr)
        for col_idx, header_text in header_cells[hr]:
            try:
                date_key, display_header, start_label, _ = _parse_header_cell(header_text, year)
            except Exception:
                continue
            display_headers_by_date[date_key] = display_header
            names: List[str] = []
            for r in name_rows:
                val = str(df.iat[r, col_idx]).strip()
                if not val:
                    continue
                if re.fullmatch(r"[-–]+", val):
                    continue
                names.append(val)
            if date_key not in timesheets[start_label]:
                timesheets[start_label][date_key] = []
            timesheets[start_label][date_key].extend(names)

    return timesheets, display_headers_by_date


def _date_sort_key(date_key: str) -> datetime:
    # date_key: "Mon 10/27/2025"
    _, mdY = date_key.split(" ", 1)
    return datetime.strptime(mdY, "%m/%d/%Y")


def _sanitize_sheet(name: str) -> str:
    # Excel forbids : \ / ? * [ ]
    return name.replace(":", "")


# ────────────────────────────────────────────────────────────────────────────────
# Core conversion
# ────────────────────────────────────────────────────────────────────────────────
def maybe_convert_adcom_excel(
    file_like,
    example_file_like: Optional[io.BytesIO] = None,
    default_year: int = 2025,
) -> bytes:
    """
    If the workbook looks like the original 'matrix' Adcom (multi-line headers),
    convert it into the example-conformant, time-sheeted format (A1:I1 & A3:I3 merged,
    9 date columns, 5 sheets). Otherwise, return the original bytes.

    Returns: Excel bytes.
    """
    # Load raw (no header inference)
    xls_bytes = file_like.getvalue() if hasattr(file_like, "getvalue") else file_like.read()
    df = pd.read_excel(io.BytesIO(xls_bytes), sheet_name=0, header=None, dtype=str).fillna("")

    # Detect
    is_matrix, header_rows, header_cells = _detect_matrix_sheet(df)
    if not is_matrix:
        # Already in a parsed/tabular or expected format — pass through
        return xls_bytes

    # Guess year from filename if available, else default
    year = default_year

    # Build timesheets
    timesheets, display_headers_by_date = _build_timesheets(df, header_rows, header_cells, year)

    # Convert 10:00 AM → 10:30 AM
    if "10:00 AM" in timesheets:
        if "10:30 AM" not in timesheets:
            timesheets["10:30 AM"] = timesheets["10:00 AM"]
        else:
            for dkey, names in timesheets["10:00 AM"].items():
                timesheets["10:30 AM"].setdefault(dkey, [])
                timesheets["10:30 AM"][dkey].extend(names)
        del timesheets["10:00 AM"]

    # Determine date headers in the final workbook (use example if provided; else derive)
    if example_file_like:
        wb_ex = load_workbook(example_file_like)
        ex_sheet = wb_ex.active
        example_banner = ex_sheet["A1"].value or ""
        example_headers = [ex_sheet.cell(row=4, column=c).value for c in range(1, 10)]
        # "Monday (10/27)" → "Mon 10/27/2025"
        dow_to_abbrev = {
            "Monday": "Mon", "Tuesday": "Tue", "Wednesday": "Wed", "Thursday": "Thu",
            "Friday": "Fri", "Saturday": "Sat", "Sunday": "Sun",
        }
        example_date_keys = []
        for h in example_headers:
            dow, rest = h.split("(", 1)
            dow = dow.strip()
            mm, dd = rest.strip(") ").split("/")
            example_date_keys.append(f"{dow_to_abbrev.get(dow, dow[:3])} {int(mm)}/{int(dd)}/{year}")
    else:
        # Derive headers from detected dates (chronological, take first 9)
        all_dates = sorted(display_headers_by_date.keys(), key=_date_sort_key)
        picked = all_dates[:9]
        example_date_keys = picked
        example_headers = [display_headers_by_date[d] for d in picked]
        # Simple banner
        example_banner = "Round 1 TBD Dates: " + " ; ".join(example_headers[:5])

    # Time windows
    time_windows = {
        "8:00 AM": ("8:00 AM", "9:30 AM"),
        "10:30 AM": ("10:30 AM", "12:00 PM"),
        "1:00 PM": ("1:00 PM", "2:30 PM"),
        "3:30 PM": ("3:30 PM", "5:00 PM"),
        "6:00 PM": ("6:00 PM", "7:30 PM"),
    }

    desired_sheets = [
        ("8 AM", "8:00 AM"),
        ("10:30 AM", "10:30 AM"),
        ("1:00 PM", "1:00 PM"),
        ("3:30 PM", "3:30 PM"),
        ("6:00 PM", "6:00 PM"),
    ]

    # Build formatted workbook into bytes
    import xlsxwriter
    out = io.BytesIO()
    wb = xlsxwriter.Workbook(out)

    fmt_title = wb.add_format({"bold": True, "align": "center", "valign": "vcenter"})
    fmt_subtitle = wb.add_format({"bold": True, "align": "center", "valign": "vcenter"})
    fmt_header = wb.add_format({"bold": True, "align": "center"})
    fmt_name = wb.add_format({"align": "left"})

    for display_name, time_key in desired_sheets:
        ws = wb.add_worksheet(_sanitize_sheet(display_name))
        # Merged title & window rows
        ws.merge_range("A1:I1", example_banner, fmt_title)
        start, end = time_windows.get(time_key, (time_key, ""))
        ws.merge_range("A3:I3", f"{start} - {end}".strip(), fmt_subtitle)
        # Row 4: headers
        for c, hdr in enumerate(example_headers):
            ws.write(3, c, hdr, fmt_header)
        # Names under each date col
        date_map = timesheets.get(time_key, OrderedDict())
        for c, dkey in enumerate(example_date_keys):
            names = list(date_map.get(dkey, []))
            for r, name in enumerate(names, start=4):
                ws.write(r, c, name, fmt_name)
        ws.set_column(0, 8, 20)

    wb.close()
    out.seek(0)
    return out.read()
